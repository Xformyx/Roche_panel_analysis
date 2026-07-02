"""
Roche_nxt Web UI - Nextflow Pipeline Manager
Order-based workflow with Dashboard, Orders, and Setup.
"""

import os
import re
import json
import glob
import shutil
import uuid
import csv
import sqlite3
import subprocess
import time
import threading
import psutil
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, g, make_response, session, Response, abort
from werkzeug.security import check_password_hash, generate_password_hash

app = Flask(__name__)

# ── Interactive Dash App Integration ──────────────────────────────────────────
# Dash is mounted at /interactive/ — disabled by default to prevent Flask conflicts.
# Enable by setting env var ENABLE_DASH=1.
if os.environ.get("ENABLE_DASH", "0") == "1":
    try:
        import interactive_app
        dash_app = interactive_app.create_dash_app(app)
        print("Interactive Dash app mounted at /interactive/")
    except ImportError as e:
        print(f"Warning: Could not initialize interactive Dash app: {e}")
    except Exception as e:
        print(f"Warning: Error initializing interactive Dash app: {e}")
# Host-mounted templates (docker) must be picked up without restarting the process
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

# ---------------------------------------------------------------------------
# Path resolution
# ---------------------------------------------------------------------------
BASE_DIR = os.environ.get("ROCHE_NXT_DIR", "/roche_nxt")
DATA_DIR = os.path.join(BASE_DIR, "data")
FASTQ_DIR = os.path.join(BASE_DIR, "fastq")
FASTQ_SOURCE_DIR = "/fastq_source"
FASTQ_HOST_DIR = os.environ.get("FASTQ_HOST_DIR", "")
BED_SOURCE_DIR = "/bed_source"
BED_HOST_DIR = os.environ.get("BED_HOST_DIR", "")
RESULTS_DIR = os.path.join(BASE_DIR, "results")
LOG_DIR = os.path.join(BASE_DIR, "log")
DB_FILE = os.path.join(BASE_DIR, "log", "orders_nxt.db")


def _get_flask_secret_key():
    key = os.environ.get("FLASK_SECRET_KEY", "").strip()
    if key:
        return key
    sk_path = os.path.join(os.path.dirname(DB_FILE), ".flask_secret")
    try:
        if os.path.isfile(sk_path):
            with open(sk_path, "r") as f:
                return f.read().strip()
        os.makedirs(os.path.dirname(sk_path), exist_ok=True)
        key = os.urandom(32).hex()
        with open(sk_path, "w") as f:
            f.write(key)
        os.chmod(sk_path, 0o600)
        return key
    except Exception:
        return "dev-insecure-change-me"


app.secret_key = _get_flask_secret_key()

# ---------------------------------------------------------------------------
# Application version  (loaded from version.json alongside app.py)
# ---------------------------------------------------------------------------
def _load_app_version() -> str:
    try:
        _vpath = os.path.join(os.path.dirname(__file__), "version.json")
        with open(_vpath, "r", encoding="utf-8") as _vf:
            return json.load(_vf).get("version", "")
    except Exception:
        return ""

APP_VERSION = _load_app_version()

HOST_DIR = os.environ.get("HOST_DIR", "/home/ken/Roche_nxt")
ANALYSIS_IMAGE = os.environ.get("ANALYSIS_IMAGE", "roche_nxt_analysis:latest")
LIFTOVER_CHAIN_HG38_TO_HG19 = os.environ.get(
    "LIFTOVER_CHAIN_HG38_TO_HG19",
    "/liftover/hg38ToHg19.over.chain.gz",
)

# ---------------------------------------------------------------------------
# Feature flags — loaded from a signed license.json (DEV_MODE bypasses this).
# ---------------------------------------------------------------------------
import license as _license  # local module: web_ui/license.py

try:
    _LICENSE_INFO = _license.load()
except _license.LicenseError as _exc:
    # Fail fast: the server cannot start without a valid license in prod.
    # In DEV_MODE this branch is never reached (license.load() always succeeds).
    import sys as _sys
    print("=" * 70, file=_sys.stderr)
    print("Roche_nxt license error: {}".format(_exc), file=_sys.stderr)
    print("  - Place a signed license at {}.".format(_license.LICENSE_PATH), file=_sys.stderr)
    print("  - Or set DEV_MODE=1 for internal development use.", file=_sys.stderr)
    print("=" * 70, file=_sys.stderr)
    raise SystemExit(2)

FEATURES = dict(_LICENSE_INFO["features"])  # copy; app.py never mutates it
LICENSE_META = {
    "customer": _LICENSE_INFO.get("customer", ""),
    "issued": _LICENSE_INFO.get("issued", ""),
    "expires": _LICENSE_INFO.get("expires", ""),
    "dev_mode": bool(_LICENSE_INFO.get("dev_mode")),
}

# Backwards-compatible aliases so existing call sites keep working.
# IGV is always a base feature; Longitudinal UI/API is gated by settings
# (longitudinal_enabled), but pipeline can still run existing L jobs.
ENABLE_LONGITUDINAL = True
ENABLE_IGV = True
ENABLE_HG19_VIEW = FEATURES["hg19_view"]

# ---------------------------------------------------------------------------
# Resource limits  (0 = auto-detect)
# ---------------------------------------------------------------------------
ESTIMATED_CPUS_PER_SAMPLE = 20
ESTIMATED_MEM_GB_PER_SAMPLE = 30

def detect_system_resources():
    """Return (total_cpus, total_memory_gb) of the host."""
    cpus = psutil.cpu_count(logical=True) or 4
    mem_gb = int(psutil.virtual_memory().total / (1024 ** 3))
    return cpus, mem_gb


def get_resource_limits():
    """Resolve effective max_cpus, max_memory, max_concurrent from env/settings/auto."""
    env_cpus = int(os.environ.get("MAX_CPUS", "0"))
    env_mem = int(os.environ.get("MAX_MEMORY", "0"))
    env_conc = int(os.environ.get("MAX_CONCURRENT_SAMPLES", "0"))

    sys_cpus, sys_mem = detect_system_resources()

    max_cpus = env_cpus if env_cpus > 0 else sys_cpus
    max_mem = env_mem if env_mem > 0 else sys_mem

    if env_conc > 0:
        max_conc = env_conc
    else:
        conc_by_cpu = max(1, max_cpus // ESTIMATED_CPUS_PER_SAMPLE)
        conc_by_mem = max(1, max_mem // ESTIMATED_MEM_GB_PER_SAMPLE)
        max_conc = min(conc_by_cpu, conc_by_mem)

    return max_cpus, max_mem, max_conc

# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------
def get_db():
    if "db" not in g:
        os.makedirs(os.path.dirname(DB_FILE), exist_ok=True)
        g.db = sqlite3.connect(DB_FILE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


@app.before_request
def _require_auth():
    ep = request.endpoint
    if ep == "static" or (request.path or "").startswith("/static/"):
        return None
    if ep in ("index", "favicon", "api_explorer"):
        return None
    if ep == "auth_me":
        return None
    if ep == "auth_login" and request.method == "POST":
        return None
    # Allow API key authentication (for external CLI / LIS integration)
    if _api_key_auth():
        return None

    uid = session.get("user_id")
    if not uid:
        return jsonify({"error": "Unauthorized", "code": "auth_required"}), 401
    urow = get_user_by_id(uid)
    if not urow:
        session.clear()
        return jsonify({"error": "Unauthorized", "code": "auth_required"}), 401
    if urow.get("must_change_password"):
        if ep in ("auth_change_password", "auth_logout", "auth_me"):
            return None
        return jsonify({"error": "Password change required", "code": "must_change_password"}), 403
    return None


def init_db():
    os.makedirs(os.path.dirname(DB_FILE), exist_ok=True)
    conn = sqlite3.connect(DB_FILE)
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS orders (
        id              TEXT PRIMARY KEY,
        order_name      TEXT NOT NULL,
        patient_name    TEXT DEFAULT '',
        patient_dob     TEXT DEFAULT '',
        chart_number    TEXT DEFAULT '',
        department      TEXT DEFAULT '',
        doctor_name     TEXT DEFAULT '',
        diagnosis       TEXT DEFAULT '',
        doctor_comment  TEXT DEFAULT '',
        sample_name     TEXT NOT NULL,
        r1_fastq        TEXT NOT NULL,
        r2_fastq        TEXT NOT NULL,
        reference       TEXT DEFAULT 'hg38',
        profile         TEXT DEFAULT 'docker',
        af_threshold    REAL DEFAULT 0.005,
        bed_file        TEXT DEFAULT '',
        delete_intermediate TEXT DEFAULT 'Y',
        order_type      TEXT DEFAULT 'baseline',
        baseline_order_id   TEXT DEFAULT '',
        germline_order_id   TEXT DEFAULT '',
        followup_order_ids  TEXT DEFAULT '',
        status          TEXT DEFAULT 'registered',
        nf_run_name     TEXT DEFAULT '',
        nf_work_dir     TEXT DEFAULT '',
        pid             INTEGER DEFAULT 0,
        error_message   TEXT DEFAULT '',
        created_at      TEXT,
        updated_at      TEXT,
        started_at      TEXT,
        completed_at    TEXT,
        created_by_user_id   TEXT DEFAULT '',
        analysis_by_user_id  TEXT DEFAULT '',
        container_name  TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS settings (
        key   TEXT PRIMARY KEY,
        value TEXT
    );
    CREATE TABLE IF NOT EXISTS users (
        id                   TEXT PRIMARY KEY,
        password_hash        TEXT NOT NULL,
        role                 TEXT NOT NULL DEFAULT 'user',
        name                 TEXT NOT NULL DEFAULT '',
        affiliation          TEXT DEFAULT '',
        phone                TEXT DEFAULT '',
        email                TEXT DEFAULT '',
        must_change_password INTEGER NOT NULL DEFAULT 1,
        created_at           TEXT,
        created_by           TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS variant_lists (
        id          TEXT PRIMARY KEY,
        list_type   TEXT NOT NULL,   -- 'blacklist' | 'whitelist' | 'onhold'
        entry_type  TEXT NOT NULL DEFAULT 'position',  -- 'cnumber' | 'position'
        cnumber     TEXT DEFAULT '',
        chrom       TEXT DEFAULT '',
        pos         INTEGER DEFAULT 0,
        ref         TEXT DEFAULT '',
        alt         TEXT DEFAULT '',
        gene        TEXT DEFAULT '',
        note        TEXT DEFAULT '',
        created_at  TEXT,
        updated_at  TEXT
    );
    """)
    for col, coldef in [
        ("order_type", "TEXT DEFAULT 'baseline'"),
        ("baseline_order_id", "TEXT DEFAULT ''"),
        ("germline_order_id", "TEXT DEFAULT ''"),
        ("followup_order_ids", "TEXT DEFAULT ''"),
        ("reuse_work_order_id", "TEXT DEFAULT ''"),
        ("created_by_user_id", "TEXT DEFAULT ''"),
        ("analysis_by_user_id", "TEXT DEFAULT ''"),
        ("subsample_info", "TEXT DEFAULT ''"),
        ("use_umi",     "TEXT DEFAULT ''"),
        ("panel_type",      "TEXT DEFAULT 'exome'"),
        ("container_name",  "TEXT DEFAULT ''"),
    ]:
        try:
            conn.execute(f"ALTER TABLE orders ADD COLUMN {col} {coldef}")
        except sqlite3.OperationalError:
            pass

    # variant_lists migrations
    for col, coldef in [
        ("category", "TEXT DEFAULT ''"),
        ("ratio",    "TEXT DEFAULT ''"),
        ("genome",   "TEXT DEFAULT 'hg38'"),   # 'hg38' | 'hg19' | 'any'
    ]:
        try:
            conn.execute(f"ALTER TABLE variant_lists ADD COLUMN {col} {coldef}")
        except sqlite3.OperationalError:
            pass
    # Back-fill genome='any' for cnumber and gene entries (reference-independent)
    conn.execute(
        "UPDATE variant_lists SET genome='any' WHERE entry_type IN ('cnumber','gene') AND (genome IS NULL OR genome='hg38')"
    )

    # Legacy: unused columns / settings keys removed from UI
    try:
        conn.execute("ALTER TABLE orders DROP COLUMN remove_bams")
    except sqlite3.OperationalError:
        pass
    conn.execute(
        "DELETE FROM settings WHERE key IN ('nf_resume', 'remove_bam', 'remove_bams')"
    )

    if not conn.execute("SELECT 1 FROM users WHERE id='admin'").fetchone():
        now_u = datetime.now().isoformat()
        conn.execute(
            """INSERT INTO users (id, password_hash, role, name, must_change_password, created_at, created_by)
               VALUES (?,?,?,?,?,?,?)""",
            (
                "admin",
                generate_password_hash("admin1234"),
                "admin",
                "Administrator",
                1,
                now_u,
                "system",
            ),
        )

    # Recovery: set admin password to admin1234 (one-time). Remove env after use.
    if os.environ.get("ROCHE_NXT_RESET_ADMIN_PASSWORD", "").lower() in ("1", "true", "yes"):
        conn.execute(
            "UPDATE users SET password_hash=?, must_change_password=1 WHERE id='admin'",
            (generate_password_hash("admin1234"),),
        )

    cur = conn.execute("SELECT COUNT(*) FROM settings")
    if cur.fetchone()[0] == 0:
        defaults = {
            "max_concurrent_samples": "3",
            "default_af_threshold": "0.005",
            "default_reference": "hg38",
            "default_profile": "docker",
            "delete_intermediate": "Y",
            "fastq_base_dir": "",
        }
        conn.executemany(
            "INSERT INTO settings (key, value) VALUES (?, ?)",
            defaults.items(),
        )
    conn.commit()
    conn.close()


def dict_from_row(row):
    return dict(row) if row else None


def get_user_by_id(user_id):
    if not user_id:
        return None
    db = get_db()
    row = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    return dict_from_row(row)


def find_user_for_login(login_id):
    """Resolve user by id case-insensitively (e.g. Admin vs admin)."""
    if not login_id:
        return None
    db = get_db()
    row = db.execute("SELECT * FROM users WHERE LOWER(id)=LOWER(?)", (login_id.strip(),)).fetchone()
    return dict_from_row(row)


def is_admin_user(user_id):
    row = get_user_by_id(user_id)
    return bool(row and row.get("role") == "admin")


def format_order_snapshot(order):
    """Human-readable lines (Korean labels) for order registration fields."""
    o = dict_from_row(order) if order is not None and not isinstance(order, dict) else order
    if not o:
        return ""

    def val(k, default="-"):
        v = o.get(k)
        if v is None or v == "":
            return default
        return str(v)

    lines = [
        f"오더명: {val('order_name')}",
        f"환자명: {val('patient_name')}",
        f"생년월일: {val('patient_dob')}",
        f"차트번호: {val('chart_number')}",
        f"진료과: {val('department')}",
        f"담당의: {val('doctor_name')}",
        f"진단: {val('diagnosis')}",
        f"의사 코멘트: {val('doctor_comment')}",
        f"샘플명: {val('sample_name')}",
        f"R1 FASTQ: {val('r1_fastq')}",
        f"R2 FASTQ: {val('r2_fastq')}",
        f"레퍼런스: {val('reference')}",
        f"AF threshold: {val('af_threshold')}",
        f"BED 파일: {val('bed_file')}",
        f"중간파일 삭제(work 정리): {val('delete_intermediate')}",
    ]
    if (o.get("order_type") or "").strip().lower() == "longitudinal":
        lines.extend(
            [
                f"Baseline 오더 ID: {val('baseline_order_id')}",
                f"Germline 오더 ID: {val('germline_order_id')}",
                f"Follow-up 오더 ID: {val('followup_order_ids')}",
            ]
        )
    lines.append(f"오더 생성일시: {val('created_at')}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def get_all_settings():
    db = get_db()
    rows = db.execute("SELECT key, value FROM settings").fetchall()
    return {r["key"]: r["value"] for r in rows}


def get_setting(key, default=""):
    db = get_db()
    row = db.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return row["value"] if row else default


def longitudinal_feature_enabled():
    """Settings → 일반: Longitudinal 분석 사용 여부 (신규 L 오더·변환·API)."""
    lo = get_all_settings().get("longitudinal_enabled", "true")
    return str(lo).lower() in ("true", "1", "yes", "on")


def get_or_create_api_key() -> str:
    """Return the stored API key, creating one if it doesn't exist yet."""
    db = get_db()
    row = db.execute("SELECT value FROM settings WHERE key='api_key'").fetchone()
    if row and row["value"]:
        return row["value"]
    key = "rnxt-" + uuid.uuid4().hex
    db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('api_key', ?)", (key,))
    db.commit()
    return key


def _api_key_auth() -> bool:
    """Return True if the request carries a valid API key."""
    stored = get_setting("api_key", "")
    if not stored:
        return False
    # Accept X-Api-Key header or Authorization: Bearer <key>
    via_header = request.headers.get("X-Api-Key", "")
    via_bearer = ""
    auth_hdr = request.headers.get("Authorization", "")
    if auth_hdr.lower().startswith("bearer "):
        via_bearer = auth_hdr[7:].strip()
    candidate = via_header or via_bearer
    return bool(candidate) and candidate == stored


def docker_container_status(container_name, retries=3):
    """Check Docker container status. Returns 'running', 'exited', or None."""
    delay = 0.2
    for attempt in range(retries):
        try:
            r = subprocess.run(
                ["docker", "inspect", "-f", "{{.State.Status}}", container_name],
                capture_output=True, text=True, timeout=10,
            )
            if r.returncode == 0:
                return r.stdout.strip()
        except Exception:
            pass
        if attempt < retries - 1:
            time.sleep(delay)
    return None


def docker_container_exit_code(container_name):
    """Get container exit code. Returns int or None."""
    try:
        r = subprocess.run(
            ["docker", "inspect", "-f", "{{.State.ExitCode}}", container_name],
            capture_output=True, text=True, timeout=10,
        )
        if r.returncode == 0:
            return int(r.stdout.strip())
    except Exception:
        pass
    return None


def _cleanup_nf_lock(order_id):
    """Remove Nextflow session LOCK files left behind by abrupt container termination."""
    nxf_home = os.path.join(BASE_DIR, "work", ".nxf_home", order_id)
    pattern = os.path.join(nxf_home, ".nextflow", "cache", "*", "db", "LOCK")
    for lock_file in glob.glob(pattern):
        try:
            os.remove(lock_file)
        except Exception:
            pass


def _nf_log_succeeded(order_id, sample_name):
    """Check Nextflow log to confirm successful pipeline completion.
    Checks both the actual .nextflow.log (in NXF_HOME) and the Docker-cmd log.
    Returns True only when a success marker is found."""
    success_markers = ["Workflow complete", "Workflow completed", "completed successfully"]
    error_markers   = ["ERROR ~", "Script compilation failed"]

    # Primary: the real .nextflow.log written by Nextflow inside NXF_HOME
    nxf_log = os.path.join(BASE_DIR, "work", ".nxf_home", order_id, ".nextflow.log")
    # Fallback: the docker-cmd log (older format; usually only contains the run command)
    cmd_log = os.path.join(LOG_DIR, f"{sample_name}_{order_id}_nf.log")

    for log_path in [nxf_log, cmd_log]:
        if not os.path.isfile(log_path):
            continue
        try:
            with open(log_path, "r", errors="replace") as f:
                content = f.read()
            if any(m in content for m in success_markers):
                return True
            if any(m in content for m in error_markers):
                return False
        except Exception:
            continue
    return False


def _archive_and_remove_container(order_id, sample_name, container_name):
    """Flush docker logs → log file, then remove the stopped container.

    This keeps the log readable via the file-based fallback while preventing
    exited containers from accumulating indefinitely.
    """
    log_path = os.path.join(LOG_DIR, f"{sample_name}_{order_id}_nf.log")
    try:
        r = subprocess.run(
            ["docker", "logs", container_name],
            capture_output=True, text=True, timeout=30,
        )
        output = (r.stdout or "") + (r.stderr or "")
        if output:
            with open(log_path, "a", errors="replace") as f:
                f.write(output)
    except Exception:
        pass
    try:
        subprocess.run(["docker", "rm", container_name],
                       capture_output=True, timeout=15)
    except Exception:
        pass


def _resolve_exited_status(db, order_id, sample_name, container_name, now, clear_error=False):
    """Determine completed vs failed for an exited container, then archive and remove it."""
    exit_code = docker_container_exit_code(container_name)
    if exit_code == 0 and _nf_log_succeeded(order_id, sample_name):
        extra = ", error_message=''" if clear_error else ""
        db.execute(
            f"UPDATE orders SET status='completed'{extra}, completed_at=?, updated_at=? WHERE id=?",
            (now, now, order_id),
        )
    else:
        reason = f"Container exited with code {exit_code}"
        if exit_code == 0:
            reason = "Container exited normally but pipeline incomplete (killed?)"
        db.execute(
            "UPDATE orders SET status='failed', error_message=?, updated_at=? WHERE id=?",
            (reason, now, order_id),
        )
    _archive_and_remove_container(order_id, sample_name, container_name)


def sync_order_statuses(db=None):
    """Reconcile DB order statuses with Docker container states.
    Accepts an optional db connection; if None, uses the request-scoped one.
    """
    own_db = False
    if db is None:
        db = get_db()
    now = datetime.now().isoformat()

    running_orders = db.execute(
        "SELECT id, sample_name, nf_run_name, container_name FROM orders WHERE status IN ('running', 'queued')"
    ).fetchall()
    for order in running_orders:
        # Use stored container_name if available; fall back to computed name
        container_name = (order["container_name"] or "").strip() \
            or f"nxt_{order['sample_name']}_{order['id'][:14]}"
        status = docker_container_status(container_name)
        if status == "running":
            continue
        elif status == "exited":
            _resolve_exited_status(db, order["id"], order["sample_name"], container_name, now)
        elif status is None:
            try:
                r = subprocess.run(
                    ["docker", "inspect", container_name],
                    capture_output=True, text=True, timeout=10,
                )
                err = (r.stderr or "") + (r.stdout or "")
                if r.returncode != 0 and "No such object" in err:
                    _cleanup_nf_lock(order["id"])
                    db.execute(
                        "UPDATE orders SET status='failed', error_message='Container not found', updated_at=? WHERE id=?",
                        (now, order["id"]),
                    )
            except Exception:
                pass

    # Recover false negatives: container was briefly missing but may have reappeared
    recovered = db.execute(
        """SELECT id, sample_name, container_name FROM orders
           WHERE status='failed' AND error_message='Container not found'"""
    ).fetchall()
    for order in recovered:
        cn = (order["container_name"] or "").strip() \
            or f"nxt_{order['sample_name']}_{order['id'][:14]}"
        st = docker_container_status(cn)
        if st == "running":
            db.execute(
                "UPDATE orders SET status='running', error_message='', updated_at=? WHERE id=?",
                (now, order["id"]),
            )
        elif st == "exited":
            _resolve_exited_status(db, order["id"], order["sample_name"], cn, now, clear_error=True)

    # Recover any 'failed' orders where container exited 0 and NF log shows success
    # (handles the case where status was wrongly set before success detection was fixed)
    possibly_succeeded = db.execute(
        """SELECT id, sample_name, container_name FROM orders
           WHERE status='failed' AND (error_message='' OR error_message IS NULL)"""
    ).fetchall()
    for order in possibly_succeeded:
        cn = (order["container_name"] or "").strip() \
            or f"nxt_{order['sample_name']}_{order['id'][:14]}"
        if docker_container_exit_code(cn) == 0 and _nf_log_succeeded(order["id"], order["sample_name"]):
            db.execute(
                "UPDATE orders SET status='completed', updated_at=? WHERE id=?",
                (now, order["id"]),
            )

    db.commit()
    if own_db:
        db.close()


def _background_sync_loop():
    """Background thread: sync order statuses every 30 s without a Flask request context."""
    while True:
        time.sleep(30)
        try:
            conn = sqlite3.connect(DB_FILE)
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA journal_mode=WAL")
            sync_order_statuses(db=conn)
            conn.close()
        except Exception:
            pass


# Start background sync thread once (daemon so it exits when the main process does)
_sync_thread = threading.Thread(target=_background_sync_loop, daemon=True, name="order-sync")
_sync_thread.start()


# ---------------------------------------------------------------------------
# Nextflow execution
# ---------------------------------------------------------------------------
def prepare_fastq_symlinks(order):
    """Create order-specific directory with symlinks to source FASTQ files."""
    order_id = order["id"]
    order_fastq_dir = os.path.join(FASTQ_DIR, order_id)
    os.makedirs(order_fastq_dir, exist_ok=True)

    source_base = FASTQ_SOURCE_DIR if os.path.isdir(FASTQ_SOURCE_DIR) else FASTQ_DIR
    r1_source = os.path.join(source_base, order["r1_fastq"])
    r2_source = os.path.join(source_base, order["r2_fastq"])
    r1_link = os.path.join(order_fastq_dir, os.path.basename(order["r1_fastq"]))
    r2_link = os.path.join(order_fastq_dir, os.path.basename(order["r2_fastq"]))

    for src, dst in [(r1_source, r1_link), (r2_source, r2_link)]:
        if os.path.lexists(dst):
            os.remove(dst)
        os.symlink(os.path.abspath(src), dst)

    return r1_link, r2_link


def generate_samplesheet(order, r1_path, r2_path):
    """Create a single-sample CSV samplesheet for Nextflow."""
    ss_dir = os.path.join(LOG_DIR, "samplesheets")
    os.makedirs(ss_dir, exist_ok=True)
    ss_path = os.path.join(ss_dir, f"{order['sample_name']}_{order['id']}.csv")
    with open(ss_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["sample_id", "fastq_1", "fastq_2"])
        writer.writerow([order["sample_name"], r1_path, r2_path])
    return ss_path


def purge_order_disk_assets(order):
    """Remove on-disk artifacts for an order: work, NXF_HOME, results, FASTQ symlink dir, logs, samplesheet."""
    order_id = order["id"]
    sample = order["sample_name"]
    reuse_wid = (order.get("reuse_work_order_id") or "").strip()

    dir_candidates = [
        os.path.join(BASE_DIR, "work", order_id),
        os.path.join(BASE_DIR, "work", ".nxf_home", order_id),
        os.path.join(FASTQ_DIR, order_id),
    ]

    # Guard: only delete results/{sample} when no other order shares the same sample_name.
    # This protects a completed baseline/followup's results when a Longitudinal companion
    # order (Case 2 promote, or reuse-mode Case 1 where sample_name was inherited) is deleted.
    def _other_order_uses_sample(sample_name, exclude_id):
        try:
            db = get_db()
            row = db.execute(
                "SELECT id FROM orders WHERE sample_name=? AND id!=? LIMIT 1",
                (sample_name, exclude_id),
            ).fetchone()
            return row is not None
        except Exception:
            return True  # err on the safe side

    can_delete_results = (
        not reuse_wid
        and not _other_order_uses_sample(sample, order_id)
    )
    if can_delete_results:
        dir_candidates.append(os.path.join(RESULTS_DIR, sample))
    nfwd = (order.get("nf_work_dir") or "").strip()
    if nfwd and can_delete_results:
        dir_candidates.append(nfwd)

    seen_real = set()
    for d in dir_candidates:
        if not d:
            continue
        try:
            if not os.path.isdir(d):
                continue
            real = os.path.realpath(d)
            if real in seen_real:
                continue
            seen_real.add(real)
            shutil.rmtree(d, ignore_errors=True)
        except OSError:
            pass

    for fpath in (
        os.path.join(LOG_DIR, "samplesheets", f"{sample}_{order_id}.csv"),
        os.path.join(LOG_DIR, f"{sample}_{order_id}_nf.log"),
    ):
        try:
            if os.path.isfile(fpath):
                os.remove(fpath)
        except OSError:
            pass


def start_analysis(order, force=False, resume=True, started_by_user_id=None, extra_nf_params=None):
    """Start a Nextflow analysis via docker run roche_nxt_analysis."""
    db = get_db()
    now = datetime.now().isoformat()
    sample = order["sample_name"]
    order_id = order["id"]

    max_cpus, max_mem, max_conc = get_resource_limits()

    running_count = db.execute(
        "SELECT COUNT(*) c FROM orders WHERE status IN ('running', 'queued')"
    ).fetchone()["c"]
    if running_count >= max_conc and not force:
        raise RuntimeError(
            f"동시 실행 제한 초과: 현재 {running_count}개 실행 중 (최대 {max_conc}개). "
            f"서버 리소스: {max_cpus} CPU, {max_mem} GB 메모리"
        )

    cpus_per_sample = max_cpus // max(1, max_conc)
    mem_per_sample = max_mem // max(1, max_conc)

    container_name = f"nxt_{sample}_{order_id[:14]}"

    if force:
        subprocess.run(["docker", "rm", "-f", container_name], capture_output=True, timeout=10)
    else:
        existing = subprocess.run(
            ["docker", "ps", "-a", "-q", "-f", f"name=^{container_name}$"],
            capture_output=True, text=True, timeout=10,
        )
        if existing.stdout.strip():
            subprocess.run(["docker", "rm", "-f", container_name], capture_output=True, timeout=10)

    host_root = HOST_DIR
    fastq_host_dir = FASTQ_HOST_DIR or os.path.join(host_root, "fastq")
    # DATA_HOST_DIR can point to a separate disk / NAS for reference data.
    # Defaults to HOST_DIR/data so existing single-server setups work unchanged.
    data_host_dir = os.environ.get("DATA_HOST_DIR", "") or os.path.join(host_root, "data")
    bed_host_dir = BED_HOST_DIR or os.path.join(data_host_dir, "bed")

    host_samplesheet_dir = os.path.join(host_root, "log", "samplesheets")
    os.makedirs(os.path.join(LOG_DIR, "samplesheets"), exist_ok=True)

    reuse_wid = (order.get("reuse_work_order_id") or "").strip()
    longitudinal_reuse = bool(
        reuse_wid and order.get("order_type") == "longitudinal"
    )

    # Longitudinal + reuse: must use the *same* --input path and NXF_HOME as the completed
    # run. A copy named S5_<new_order>.csv breaks Nextflow cache keys; a fresh NXF_HOME has
    # no session history. Share work/{reuse_id} and work/.nxf_home/{reuse_id} with S5.
    if longitudinal_reuse:
        ss_name = f"{sample}_{reuse_wid}.csv"
        ss_host_path = os.path.join(LOG_DIR, "samplesheets", ss_name)
        ss_container_path = f"/work_nxt/log/samplesheets/{ss_name}"
        if not os.path.isfile(ss_host_path):
            r1_container = f"/work_nxt_fastq_source/{order['r1_fastq']}"
            r2_container = f"/work_nxt_fastq_source/{order['r2_fastq']}"
            with open(ss_host_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["sample_id", "fastq_1", "fastq_2"])
                writer.writerow([sample, r1_container, r2_container])
        work_dir_rel = f"work/{reuse_wid}"
        nxf_home_rel = f"work/.nxf_home/{reuse_wid}"
    else:
        ss_name = f"{sample}_{order_id}.csv"
        ss_host_path = os.path.join(LOG_DIR, "samplesheets", ss_name)
        ss_container_path = f"/work_nxt/log/samplesheets/{ss_name}"
        r1_container = f"/work_nxt_fastq_source/{order['r1_fastq']}"
        r2_container = f"/work_nxt_fastq_source/{order['r2_fastq']}"
        with open(ss_host_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["sample_id", "fastq_1", "fastq_2"])
            writer.writerow([sample, r1_container, r2_container])
        work_dir_rel = f"work/{order_id}"
        nxf_home_rel = f"work/.nxf_home/{order_id}"

    # Always generate a fresh timestamped run name to avoid "already used" collisions.
    # Nextflow -resume finds the cache from the work-dir, not from the run name.
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    run_name = f"run_{sample}_{order_id[:14]}_{ts}"

    os.makedirs(os.path.join(BASE_DIR, work_dir_rel), exist_ok=True)
    os.makedirs(os.path.join(BASE_DIR, nxf_home_rel), exist_ok=True)
    os.makedirs(os.path.join(RESULTS_DIR, sample), exist_ok=True)
    os.makedirs(LOG_DIR, exist_ok=True)

    af = order.get("af_threshold") or 0.005
    ref = order.get("reference") or "hg38"
    if ref not in ("hg38", "hg19"):
        ref = "hg38"

    # Read baseline params from settings DB (fallback to nextflow.config defaults)
    settings = get_all_settings()
    bl_seqtk_size    = settings.get("seqtk_sample_size",    "40000000")
    bl_seqtk_seed    = settings.get("seqtk_seed",           "12345")
    bl_fastp_options = settings.get("fastp_options",         "-g -W 5 -q 20 -u 40 -x -3 -l 75 -c")
    bl_min_reads     = settings.get("min_reads",             "1")
    bl_min_bq        = settings.get("min_base_quality",      "20")
    bl_max_rer       = settings.get("max_read_error_rate",   "0.025")
    bl_max_ber       = settings.get("max_base_error_rate",   "0.1")
    bl_max_ncf       = settings.get("max_no_call_fraction",  "0.1")

    # ── Subsampling: check FASTQ file sizes ──────────────────────────────────
    enable_sub = settings.get("enable_subsampling", "false").lower() in ("true", "1", "yes")
    sub_threshold_gb = float(settings.get("subsample_threshold_gb", "20") or "20")
    do_subsample = False
    subsample_info_dict = {}
    if enable_sub:
        fastq_base = FASTQ_SOURCE_DIR if os.path.isdir(FASTQ_SOURCE_DIR) else FASTQ_DIR
        r1_path_host = os.path.join(fastq_base, order["r1_fastq"])
        r2_path_host = os.path.join(fastq_base, order["r2_fastq"])
        try:
            r1_bytes = os.path.getsize(r1_path_host) if os.path.isfile(r1_path_host) else 0
            r2_bytes = os.path.getsize(r2_path_host) if os.path.isfile(r2_path_host) else 0
            total_gb = (r1_bytes + r2_bytes) / 1e9
            if total_gb >= sub_threshold_gb:
                do_subsample = True
                subsample_info_dict = {
                    "enabled": True,
                    "r1_gb": round(r1_bytes / 1e9, 2),
                    "r2_gb": round(r2_bytes / 1e9, 2),
                    "total_gb": round(total_gb, 2),
                    "threshold_gb": sub_threshold_gb,
                    "target_reads": int(bl_seqtk_size),
                    "seed": int(bl_seqtk_seed),
                }
        except OSError:
            pass
    # Persist subsample_info in order record
    sub_info_json = json.dumps(subsample_info_dict) if subsample_info_dict else ""
    db.execute(
        "UPDATE orders SET subsample_info=? WHERE id=?",
        (sub_info_json, order_id),
    )
    db.commit()

    uid_gid = subprocess.run(["id", "-u"], capture_output=True, text=True).stdout.strip()
    gid = subprocess.run(["id", "-g"], capture_output=True, text=True).stdout.strip()

    # UMI mode resolution: per-order override > global setting (default true)
    enable_umi_global = settings.get("enable_umi", "true").lower() in ("true", "1", "yes")
    order_use_umi = (order.get("use_umi") or "").strip().upper()
    if order_use_umi == "Y":
        use_umi_flag = True
    elif order_use_umi == "N":
        use_umi_flag = False
    else:
        use_umi_flag = enable_umi_global
    umi_structure = settings.get("umi_read_structure", "3M3S+T 3M3S+T")

    panel_type = (order.get("panel_type") or "exome").strip().lower()

    if panel_type == "rna":
        # ── RNAseq pipeline ────────────────────────────────────────────────
        # RNA-specific reference paths: use settings value if explicitly set,
        # otherwise fall back to the default container-side paths under /work_nxt_data
        _rna_data = "/work_nxt_data"
        _rna_defaults = {
            "star_index": f"{_rna_data}/refs/{ref}/star_index",
            "gtf":        f"{_rna_data}/refs/{ref}/gencode.v44.annotation.gtf",
            "bed12":      f"{_rna_data}/refs/{ref}/genes.bed12",
            "ctat_lib":   f"{_rna_data}/refs/{ref}/ctat_lib/ctat_genome_lib_build_dir",
        }
        rna_star_index = settings.get("rna_star_index", "").strip() or _rna_defaults["star_index"]
        rna_gtf        = settings.get("rna_gtf",        "").strip() or _rna_defaults["gtf"]
        rna_bed12      = settings.get("rna_bed12",      "").strip() or _rna_defaults["bed12"]
        rna_ctat_lib   = settings.get("rna_ctat_lib",   "").strip() or _rna_defaults["ctat_lib"]
        rna_fastp_opts = settings.get("rna_fastp_options", "-g -W 5 -q 20 -u 40 -x -3 -l 50 -c").strip()
        nf_cmd = [
            "nextflow", "run", "/work_nxt/workflows/rnaseq.nf",
            "-c", "/work_nxt/conf/rnaseq.config",
            "-profile", "local",
            "-name", run_name,
            "-work-dir", f"/work_nxt/{work_dir_rel}",
            "--input", ss_container_path,
            "--outdir", "/work_nxt/results",
            "--reference", ref,
            "--data_dir", "/work_nxt_data",
            "--max_cpus",   str(cpus_per_sample),
            "--max_memory", str(mem_per_sample),
            "--fastp_options", rna_fastp_opts,
            "--star_index", rna_star_index,
            "--gtf",        rna_gtf,
            "--bed12",      rna_bed12,
            "--ctat_lib",   rna_ctat_lib,
        ]
    else:
        # ── Exome (ctDNA) pipeline — existing logic ────────────────────────
        nf_cmd = [
            "nextflow", "run", "/work_nxt/main.nf",
            "-profile", "local",
            "-name", run_name,
            "-work-dir", f"/work_nxt/{work_dir_rel}",
            "--input", ss_container_path,
            "--outdir", "/work_nxt/results",
            "--reference", ref,
            "--af_threshold", str(af),
            "--data_dir", "/work_nxt_data",
            # UMI mode
            "--use_umi",            "true" if use_umi_flag else "false",
            "--umi_read_structure", umi_structure,
            # Subsampling flag (determined by file-size check above)
            "--subsample", "true" if do_subsample else "false",
            # Baseline params from settings DB
            "--seqtk_sample_size",    bl_seqtk_size,
            "--seqtk_seed",           bl_seqtk_seed,
            "--fastp_options",        bl_fastp_options,
            "--min_reads",            bl_min_reads,
            "--min_base_quality",     bl_min_bq,
            "--max_read_error_rate",  bl_max_rer,
            "--max_base_error_rate",  bl_max_ber,
            "--max_no_call_fraction", bl_max_ncf,
        ]

    if resume and not force:
        nf_cmd.append("-resume")

    if panel_type == "rna":
        # RNAseq pipeline: skip exome-specific flags
        if extra_nf_params:
            nf_cmd.extend(extra_nf_params)
    else:
        # Exome-specific options
        if order.get("bed_file"):
            nf_cmd.extend(["--target_bed", f"/work_nxt_bed/{order['bed_file']}"])
        if order.get("delete_intermediate") == "Y":
            nf_cmd.append("--delete_intermediate")

        nf_cmd.extend(["--max_cpus", str(cpus_per_sample), "--max_memory", str(mem_per_sample)])

    if panel_type != "rna" and order.get("order_type") == "longitudinal":
        nf_cmd.extend(["--run_select_reporter", "true", "--run_longitudinal", "true"])

        # Select Reporter parameters from settings DB
        settings = get_all_settings()
        nf_cmd.extend([
            "--sr_germline_cutoff", str(settings.get("sr_germline_cutoff", "0.005")),
            "--sr_min_af",          str(settings.get("sr_min_af",          "0.005")),
            "--sr_max_af",          str(settings.get("sr_max_af",          "0.35")),
            "--sr_min_dp",          str(settings.get("sr_min_dp",          "1000")),
            "--sr_min_vd",          str(settings.get("sr_min_vd",          "15")),
            "--sr_min_mq",          str(settings.get("sr_min_mq",          "55")),
            "--sr_min_qual",        str(settings.get("sr_min_qual",        "45")),
            "--sr_min_sbf",         str(settings.get("sr_min_sbf",         "1e-05")),
            "--sr_max_nm",          str(settings.get("sr_max_nm",          "4")),
            "--la_reads_threshold", str(settings.get("la_reads_threshold", "1000")),
            "--la_pvalue_threshold",str(settings.get("la_pvalue_threshold","0.001")),
            "--la_vaf_threshold",   str(settings.get("la_vaf_threshold",   "0.1")),
            "--la_n_sim",           str(settings.get("la_n_sim",           "10000")),
            "--la_blist_type",      str(settings.get("la_blist_type",      "variant")),
        ])

        # Pass Baseline VCF and Germline BAM paths so SELECT_REPORTERS uses
        # the correct cross-sample inputs (not the Followup's own data).
        baseline_id = (order.get("baseline_order_id") or "").strip()
        germline_id = (order.get("germline_order_id") or "").strip()

        if baseline_id and germline_id:
            db_conn = get_db()
            baseline_row = db_conn.execute("SELECT sample_name FROM orders WHERE id=?", (baseline_id,)).fetchone()
            germline_row = db_conn.execute("SELECT sample_name FROM orders WHERE id=?", (germline_id,)).fetchone()

            if baseline_row and germline_row:
                bl_sample = baseline_row["sample_name"]
                gm_sample = germline_row["sample_name"]

                results_rel = os.path.relpath(RESULTS_DIR, BASE_DIR)

                # Baseline annotated VCF txt
                bl_ann_host = os.path.join(RESULTS_DIR, bl_sample, "output", f"{bl_sample}_vardict_annotated_vcf.txt")
                if not os.path.isfile(bl_ann_host):
                    import glob as _glob2
                    bl_candidates = _glob2.glob(os.path.join(RESULTS_DIR, bl_sample, "output", "**", f"{bl_sample}_vardict_annotated_vcf.txt"), recursive=True)
                    if bl_candidates:
                        bl_ann_host = bl_candidates[0]
                if os.path.isfile(bl_ann_host):
                    bl_ann_container = f"/work_nxt/{os.path.relpath(bl_ann_host, BASE_DIR)}"
                    nf_cmd.extend(["--longitudinal_baseline_ann_txt", bl_ann_container])

                # Germline clipped_sorted BAM (prefer results/, fallback to work/)
                gm_bam_host = os.path.join(RESULTS_DIR, gm_sample, "output", "bam", f"{gm_sample}_clipped_sorted.bam")
                gm_bai_host = gm_bam_host + ".bai"
                if not os.path.isfile(gm_bam_host):
                    import glob as _glob3
                    gm_candidates = _glob3.glob(os.path.join(BASE_DIR, "work", "**", f"{gm_sample}_clipped_sorted.bam"), recursive=True)
                    # Prefer the actual file over symlinks
                    gm_candidates_real = [p for p in gm_candidates if not os.path.islink(p)]
                    if gm_candidates_real:
                        gm_bam_host = gm_candidates_real[0]
                        gm_bai_host = gm_bam_host[:-4] + ".bai"
                    elif gm_candidates:
                        gm_bam_host = os.path.realpath(gm_candidates[0])
                        gm_bai_host = gm_bam_host[:-4] + ".bai"
                if os.path.isfile(gm_bam_host):
                    gm_bam_container = f"/work_nxt/{os.path.relpath(gm_bam_host, BASE_DIR)}"
                    gm_bai_container = f"/work_nxt/{os.path.relpath(gm_bai_host, BASE_DIR)}"
                    nf_cmd.extend([
                        "--longitudinal_germline_bam", gm_bam_container,
                        "--longitudinal_germline_bai", gm_bai_container,
                    ])

    if panel_type != "rna" and extra_nf_params:
        nf_cmd.extend(extra_nf_params)

    docker_cmd = [
        "docker", "run", "-d",
        "--user", f"{uid_gid}:{gid}",
        "-e", f"HOME=/work_nxt/{nxf_home_rel}",
        "-e", f"NXF_HOME=/work_nxt/{nxf_home_rel}",
        "-e", "NXF_ANSI_LOG=false",
        "-w", f"/work_nxt/{nxf_home_rel}",
        "-v", f"{host_root}:/work_nxt",
        "-v", f"{fastq_host_dir}:/work_nxt_fastq_source:ro",
        "-v", f"{data_host_dir}:/work_nxt_data:ro",
        "-v", f"{bed_host_dir}:/work_nxt_bed:ro",
        "--name", container_name,
        ANALYSIS_IMAGE,
    ] + nf_cmd

    log_path = os.path.join(LOG_DIR, f"{sample}_{order_id}_nf.log")
    with open(log_path, "w") as lf:
        lf.write(f"# Docker command:\n# {' '.join(docker_cmd)}\n\n")

    result = subprocess.run(docker_cmd, capture_output=True, text=True, timeout=30)
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or "Docker run failed")

    container_id = result.stdout.strip()[:12]

    analyst = started_by_user_id if started_by_user_id else ""
    db.execute(
        "UPDATE orders SET status='running', nf_run_name=?, nf_work_dir=?, container_name=?, pid=0, started_at=?, updated_at=?, analysis_by_user_id=? WHERE id=?",
        (run_name, os.path.join(BASE_DIR, work_dir_rel), container_name, now, now, analyst, order_id),
    )
    db.commit()
    return container_id


# ---------------------------------------------------------------------------
# FASTQ / BED file browsing
# ---------------------------------------------------------------------------
def browse_fastq(subdir=""):
    base = FASTQ_SOURCE_DIR if os.path.isdir(FASTQ_SOURCE_DIR) else FASTQ_DIR
    target = os.path.join(base, subdir) if subdir else base
    target = os.path.abspath(target)
    if not target.startswith(os.path.abspath(base)):
        target, subdir = base, ""

    dirs, samples = [], {}
    if not os.path.exists(target):
        return {"directories": dirs, "samples": [], "current_path": subdir, "parent_path": None}

    for item in sorted(os.listdir(target)):
        full = os.path.join(target, item)
        if item.startswith("."):
            continue
        if os.path.isdir(full):
            dirs.append({"name": item, "path": os.path.join(subdir, item) if subdir else item})

    patterns = [
        r"(.+)_R([12])_\d+\.fastq[-\d]*\.gz$",
        r"(.+)_R([12])_\d+\.(fastq|fq)\.gz$",
        r"(.+)_R([12])_\d+\.(fastq|fq)$",
        r"(.+)_R([12])\.(fastq|fq)\.gz$",
        r"(.+)_R([12])\.(fastq|fq)$",
        r"(.+)_([12])\.(fastq|fq)\.gz$",
        r"(.+)_([12])\.(fastq|fq)$",
    ]

    fq_files = []
    for ext in ("*.fastq*.gz", "*.fq*.gz", "*.fastq", "*.fq"):
        fq_files.extend(glob.glob(os.path.join(target, ext)))

    for f in fq_files:
        bn = os.path.basename(f)
        for pat in patterns:
            m = re.match(pat, bn)
            if m:
                sname, rnum = m.group(1), m.group(2)
                samples.setdefault(sname, {})
                key = "r1" if rnum == "1" else "r2"
                samples[sname][key] = bn
                break

    pairs = []
    for sname, files in samples.items():
        if "r1" in files and "r2" in files:
            r1p = os.path.join(subdir, files["r1"]) if subdir else files["r1"]
            r2p = os.path.join(subdir, files["r2"]) if subdir else files["r2"]
            try:
                sz1 = os.path.getsize(os.path.join(target, files["r1"]))
                sz2 = os.path.getsize(os.path.join(target, files["r2"]))
            except OSError:
                sz1 = sz2 = 0
            pairs.append({"name": sname, "r1": r1p, "r2": r2p, "size_r1": sz1, "size_r2": sz2})

    return {
        "directories": dirs,
        "samples": sorted(pairs, key=lambda x: x["name"]),
        "current_path": subdir,
        "parent_path": os.path.dirname(subdir) if subdir else None,
    }


def browse_bed(subdir=""):
    base = BED_SOURCE_DIR if os.path.isdir(BED_SOURCE_DIR) else os.path.join(DATA_DIR, "bed")
    target = os.path.join(base, subdir) if subdir else base
    target = os.path.abspath(target)
    if not target.startswith(os.path.abspath(base)):
        target, subdir = base, ""

    dirs, files = [], []
    if os.path.exists(target):
        for item in sorted(os.listdir(target)):
            if item.startswith("."):
                continue
            full = os.path.join(target, item)
            rel = os.path.join(subdir, item) if subdir else item
            if os.path.isdir(full):
                dirs.append({"name": item, "path": rel})
            elif item.endswith(".bed"):
                files.append({"name": item, "path": rel, "size": os.path.getsize(full)})

    return {
        "directories": dirs,
        "files": files,
        "current_path": subdir,
        "parent_path": os.path.dirname(subdir) if subdir else None,
    }


# ---------------------------------------------------------------------------
# Routes — Auth & admin users
# ---------------------------------------------------------------------------
@app.route("/api/auth/me")
def auth_me():
    uid = session.get("user_id")
    if not uid:
        return jsonify({"authenticated": False})
    row = get_user_by_id(uid)
    if not row:
        session.clear()
        return jsonify({"authenticated": False})
    return jsonify({
        "authenticated": True,
        "user_id": row["id"],
        "name": row.get("name") or "",
        "role": row.get("role") or "user",
        "must_change_password": bool(row.get("must_change_password")),
    })


@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    data = request.json or {}
    user_id = (data.get("user_id") or data.get("username") or "").strip()
    password = data.get("password") or ""
    if not user_id or not password:
        return jsonify({"success": False, "error": "ID와 비밀번호를 입력하세요."}), 400
    row = find_user_for_login(user_id)
    if not row or not check_password_hash(row["password_hash"], password):
        return jsonify({"success": False, "error": "로그인 정보가 올바르지 않습니다."}), 401
    session["user_id"] = row["id"]
    session.permanent = True
    return jsonify({
        "success": True,
        "user_id": row["id"],
        "role": row.get("role"),
        "must_change_password": bool(row.get("must_change_password")),
    })


@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    session.clear()
    return jsonify({"success": True})


@app.route("/api/auth/api_key", methods=["GET"])
def auth_get_api_key():
    """Return (or auto-create) the API key for external CLI integration."""
    return jsonify({"api_key": get_or_create_api_key()})


@app.route("/api/auth/api_key/regenerate", methods=["POST"])
def auth_regenerate_api_key():
    """Generate a new API key, invalidating the previous one."""
    new_key = "rnxt-" + uuid.uuid4().hex
    db = get_db()
    db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('api_key', ?)", (new_key,))
    db.commit()
    return jsonify({"api_key": new_key})


@app.route("/api/auth/change-password", methods=["POST"])
def auth_change_password():
    uid = session.get("user_id")
    if not uid:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    data = request.json or {}
    current = data.get("current_password") or ""
    new_pw = data.get("new_password") or ""
    if len(new_pw) < 4:
        return jsonify({"success": False, "error": "새 비밀번호는 4자 이상이어야 합니다."}), 400
    row = get_user_by_id(uid)
    if not row or not check_password_hash(row["password_hash"], current):
        return jsonify({"success": False, "error": "현재 비밀번호가 올바르지 않습니다."}), 400
    db = get_db()
    db.execute(
        "UPDATE users SET password_hash=?, must_change_password=0 WHERE id=?",
        (generate_password_hash(new_pw), uid),
    )
    db.commit()
    return jsonify({"success": True})


@app.route("/api/admin/users")
def api_admin_users_list():
    uid = session.get("user_id")
    if not is_admin_user(uid):
        return jsonify({"success": False, "error": "Forbidden"}), 403
    db = get_db()
    rows = db.execute(
        "SELECT id, name, affiliation, phone, email, role, must_change_password, created_at, created_by "
        "FROM users ORDER BY created_at DESC"
    ).fetchall()
    return jsonify({"success": True, "users": [dict_from_row(r) for r in rows]})


DEFAULT_USER_PASSWORD = "user1234"


@app.route("/api/admin/users/<user_id>/reset-password", methods=["POST"])
def api_admin_users_reset_password(user_id):
    """Admin-only: force-reset a general user's password.

    The next login will be prompted to change the password (must_change_password=1).
    To keep the attack surface small this endpoint refuses to:
      - reset an admin's password (including the admin's own)
      - reset a non-existent user
    """
    acting_uid = session.get("user_id")
    if not is_admin_user(acting_uid):
        return jsonify({"success": False, "error": "Forbidden"}), 403

    target_id = (user_id or "").strip()
    if not target_id:
        return jsonify({"success": False, "error": "사용자 ID가 필요합니다."}), 400

    target = get_user_by_id(target_id)
    if not target:
        return jsonify({"success": False, "error": "존재하지 않는 사용자입니다."}), 404

    if (target.get("role") or "") == "admin":
        return jsonify({
            "success": False,
            "error": "관리자 계정의 비밀번호는 이 기능으로 초기화할 수 없습니다.",
        }), 403

    data = request.json or {}
    new_pw = (data.get("new_password") or "").strip() or DEFAULT_USER_PASSWORD
    if len(new_pw) < 4:
        return jsonify({"success": False, "error": "비밀번호는 4자 이상이어야 합니다."}), 400

    db = get_db()
    db.execute(
        "UPDATE users SET password_hash=?, must_change_password=1 WHERE id=?",
        (generate_password_hash(new_pw), target_id),
    )
    db.commit()

    return jsonify({
        "success": True,
        "user_id": target_id,
        "temporary_password": new_pw,
        "must_change_password": True,
    })


@app.route("/api/admin/users", methods=["POST"])
def api_admin_users_create():
    uid = session.get("user_id")
    if not is_admin_user(uid):
        return jsonify({"success": False, "error": "Forbidden"}), 403
    data = request.json or {}
    new_id = (data.get("user_id") or data.get("id") or "").strip()
    name = (data.get("name") or "").strip()
    if not new_id or not name:
        return jsonify({"success": False, "error": "사용자 ID와 이름은 필수입니다."}), 400
    db = get_db()
    if db.execute("SELECT 1 FROM users WHERE id=?", (new_id,)).fetchone():
        return jsonify({"success": False, "error": "이미 존재하는 ID입니다."}), 400
    now = datetime.now().isoformat()
    default_pw = DEFAULT_USER_PASSWORD
    db.execute(
        """INSERT INTO users (id, password_hash, role, name, affiliation, phone, email,
           must_change_password, created_at, created_by)
           VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (
            new_id,
            generate_password_hash(default_pw),
            "user",
            name,
            (data.get("affiliation") or "").strip(),
            (data.get("phone") or "").strip(),
            (data.get("email") or "").strip(),
            1,
            now,
            uid,
        ),
    )
    db.commit()
    return jsonify({"success": True, "user_id": new_id})


# ---------------------------------------------------------------------------
# Routes — Pages
# ---------------------------------------------------------------------------
@app.route("/favicon.ico")
def favicon():
    """Some browsers only request /favicon.ico; serve icon with no-cache headers."""
    path = os.path.join(app.static_folder, "leaninbio-icon.png")
    if not os.path.isfile(path):
        return ("", 404)
    resp = make_response(send_file(path, mimetype="image/png", max_age=0))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@app.after_request
def _static_no_cache_for_icons(response):
    """Prevent aggressive favicon caching when served from /static/ in dev."""
    if request.path.startswith("/static/") and "leaninbio-icon" in request.path:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return response


@app.route("/")
def index():
    return render_template(
        "index.html",
        has_auth_session=bool(session.get("user_id")),
        app_version=APP_VERSION,
    )


@app.route("/developer")
def api_explorer():
    """API Explorer — 3rd party developer guide (no auth required)."""
    return render_template("api_explorer.html", app_version=APP_VERSION)


@app.route("/api/features")
def api_features():
    chain_ready = ENABLE_HG19_VIEW and os.path.isfile(LIFTOVER_CHAIN_HG38_TO_HG19)
    return jsonify({
        "longitudinal": longitudinal_feature_enabled(),
        "igv": True,
        "hg19_view": bool(chain_ready),
        "hg19_chain_missing": ENABLE_HG19_VIEW and not chain_ready,
        "license": LICENSE_META,
    })


@app.route("/api/system_resources")
def api_system_resources():
    """Return detected system resources and current effective limits."""
    sys_cpus, sys_mem = detect_system_resources()
    max_cpus, max_mem, max_conc = get_resource_limits()
    cpus_per_sample = max_cpus // max(1, max_conc)
    mem_per_sample = max_mem // max(1, max_conc)
    db = get_db()
    running = db.execute(
        "SELECT COUNT(*) c FROM orders WHERE status IN ('running', 'queued')"
    ).fetchone()["c"]
    return jsonify({
        "system_cpus": sys_cpus,
        "system_memory_gb": sys_mem,
        "max_cpus": max_cpus,
        "max_memory_gb": max_mem,
        "max_concurrent_samples": max_conc,
        "cpus_per_sample": cpus_per_sample,
        "memory_per_sample_gb": mem_per_sample,
        "currently_running": running,
        "estimated_cpus_per_sample": ESTIMATED_CPUS_PER_SAMPLE,
        "estimated_mem_per_sample": ESTIMATED_MEM_GB_PER_SAMPLE,
    })


# ---------------------------------------------------------------------------
# Routes — Dashboard API
# ---------------------------------------------------------------------------
@app.route("/api/dashboard")
def api_dashboard():
    db = get_db()
    sync_order_statuses()
    total = db.execute("SELECT COUNT(*) c FROM orders").fetchone()["c"]
    running = db.execute("SELECT COUNT(*) c FROM orders WHERE status='running'").fetchone()["c"]
    completed = db.execute("SELECT COUNT(*) c FROM orders WHERE status='completed'").fetchone()["c"]
    failed = db.execute("SELECT COUNT(*) c FROM orders WHERE status='failed'").fetchone()["c"]
    queued = db.execute("SELECT COUNT(*) c FROM orders WHERE status='queued'").fetchone()["c"]
    registered = db.execute("SELECT COUNT(*) c FROM orders WHERE status='registered'").fetchone()["c"]

    recent = db.execute(
        "SELECT id, order_name, sample_name, status, updated_at, panel_type FROM orders ORDER BY updated_at DESC LIMIT 10"
    ).fetchall()

    return jsonify({
        "total": total, "running": running, "completed": completed,
        "failed": failed, "queued": queued, "registered": registered,
        "recent": [dict_from_row(r) for r in recent],
    })


@app.route("/api/resources")
def api_resources():
    cpu = psutil.cpu_percent(interval=0.5)
    mem = psutil.virtual_memory()

    run_path = BASE_DIR if os.path.exists(BASE_DIR) else "/"
    disk_run = psutil.disk_usage(run_path)

    fastq_path = FASTQ_HOST_DIR if FASTQ_HOST_DIR and os.path.exists(FASTQ_HOST_DIR) else run_path
    disk_fastq = psutil.disk_usage(fastq_path)

    return jsonify({
        "cpu_percent": cpu,
        "cpu_count": psutil.cpu_count(),
        "mem_total_gb": round(mem.total / 1e9, 1),
        "mem_used_gb": round(mem.used / 1e9, 1),
        "mem_percent": mem.percent,
        # 실행환경 디스크
        "disk_total_gb":   round(disk_run.total / 1e9, 1),
        "disk_used_gb":    round(disk_run.used  / 1e9, 1),
        "disk_percent":    disk_run.percent,
        # Fastq 디스크 (실행환경과 같은 파티션이면 동일 값)
        "disk_fastq_total_gb":  round(disk_fastq.total / 1e9, 1),
        "disk_fastq_used_gb":   round(disk_fastq.used  / 1e9, 1),
        "disk_fastq_percent":   disk_fastq.percent,
        "disk_fastq_same":      os.path.realpath(fastq_path) == os.path.realpath(run_path)
                                or disk_run.total == disk_fastq.total,
    })


# ---------------------------------------------------------------------------
# Routes — Orders API
# ---------------------------------------------------------------------------
@app.route("/api/orders")
def api_orders():
    db = get_db()
    sync_order_statuses()
    status_filter = request.args.get("status")
    if status_filter:
        rows = db.execute("SELECT * FROM orders WHERE status=? ORDER BY created_at DESC", (status_filter,)).fetchall()
    else:
        rows = db.execute("SELECT * FROM orders ORDER BY created_at DESC").fetchall()

    result = []
    for r in rows:
        o = dict_from_row(r)
        sample = o.get("sample_name", "")
        o["has_vcf"] = bool(glob.glob(os.path.join(RESULTS_DIR, sample, "**", f"{sample}*.vcf"), recursive=True))
        o["has_log"] = bool(glob.glob(os.path.join(LOG_DIR, f"{sample}_*_nf.log")))
        qc_dir = os.path.join(RESULTS_DIR, sample, "QC_report")
        o["has_qc"] = os.path.isdir(qc_dir) and bool(os.listdir(qc_dir))
        o["has_expression"] = bool(glob.glob(os.path.join(
            RESULTS_DIR, sample, "expression_plots", f"{sample}_expression_summary.tsv")))
        result.append(o)
    return jsonify(result)


@app.route("/api/orders/completed_list")
def api_completed_list():
    """Return completed orders that can serve as Baseline/Germline/Followup references."""
    db = get_db()
    rows = db.execute(
        "SELECT id, order_name, patient_name, sample_name, chart_number, order_type, status, created_at "
        "FROM orders WHERE status='completed' ORDER BY created_at DESC"
    ).fetchall()
    return jsonify([dict_from_row(r) for r in rows])


@app.route("/api/orders", methods=["POST"])
def api_create_order():
    data = request.json
    order_type = data.get("order_type", "baseline")
    if order_type == "longitudinal" and not longitudinal_feature_enabled():
        return jsonify({"success": False, "error": "Longitudinal 분석이 설정에서 비활성화되어 있습니다."}), 403
    reuse_wid = (data.get("reuse_work_order_id") or "").strip()
    if reuse_wid:
        if order_type != "longitudinal":
            return jsonify({"success": False, "error": "reuse_work_order_id는 Longitudinal에서만 사용할 수 있습니다."}), 400

    if order_type == "longitudinal" and reuse_wid:
        db = get_db()
        src = db.execute("SELECT * FROM orders WHERE id=?", (reuse_wid,)).fetchone()
        if not src:
            return jsonify({"success": False, "error": "reuse_work_order_id: 오더를 찾을 수 없습니다."}), 400
        src = dict_from_row(src)
        if src.get("status") != "completed":
            return jsonify({"success": False, "error": "재사용할 오더는 분석 완료(completed) 상태여야 합니다."}), 400
        sample_name = src["sample_name"]
        r1_fastq = src["r1_fastq"]
        r2_fastq = src["r2_fastq"]
    else:
        if not data.get("sample_name") or not data.get("r1_fastq") or not data.get("r2_fastq"):
            return jsonify({"success": False, "error": "sample_name, r1_fastq, r2_fastq are required"}), 400
        sample_name = data["sample_name"]
        r1_fastq = data["r1_fastq"]
        r2_fastq = data["r2_fastq"]

    order_id = datetime.now().strftime("%Y%m%d%H%M%S") + "-" + uuid.uuid4().hex[:6]
    now = datetime.now().isoformat()

    if order_type == "longitudinal":
        if not data.get("baseline_order_id") or not data.get("germline_order_id"):
            return jsonify({"success": False, "error": "Longitudinal requires baseline and germline order selection"}), 400

    followup_ids = ",".join(data.get("followup_order_ids", [])) if isinstance(data.get("followup_order_ids"), list) else data.get("followup_order_ids", "")

    db = get_db()
    created_by = session.get("user_id") or ""
    # use_umi: '' (inherit global), 'Y', or 'N'
    use_umi_val = (data.get("use_umi") or "").strip().upper()
    if use_umi_val not in ("Y", "N", ""):
        use_umi_val = ""

    # panel_type: 'exome' (default) or 'rna'
    panel_type_val = (data.get("panel_type") or "exome").strip().lower()
    if panel_type_val not in ("exome", "rna"):
        panel_type_val = "exome"

    db.execute("""
        INSERT INTO orders (id, order_name, patient_name, patient_dob, chart_number,
            department, doctor_name, diagnosis, doctor_comment,
            sample_name, r1_fastq, r2_fastq, reference, profile,
            af_threshold, bed_file, delete_intermediate,
            order_type, baseline_order_id, germline_order_id, followup_order_ids, reuse_work_order_id,
            use_umi, panel_type,
            status, created_at, updated_at, created_by_user_id)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        order_id,
        data.get("order_name", sample_name),
        data.get("patient_name", ""),
        data.get("patient_dob", ""),
        data.get("chart_number", ""),
        data.get("department", ""),
        data.get("doctor_name", ""),
        data.get("diagnosis", ""),
        data.get("doctor_comment", ""),
        sample_name,
        r1_fastq,
        r2_fastq,
        data.get("reference", "hg38") if data.get("reference") in ("hg38", "hg19") else "hg38",
        data.get("profile", "docker"),
        float(data.get("af_threshold", 0.005)),
        data.get("bed_file", ""),
        data.get("delete_intermediate", "Y"),
        order_type,
        data.get("baseline_order_id", ""),
        data.get("germline_order_id", ""),
        followup_ids,
        reuse_wid,
        use_umi_val,
        panel_type_val,
        "registered", now, now,
        created_by,
    ))
    db.commit()
    return jsonify({"success": True, "order_id": order_id})


@app.route("/api/orders/<order_id>")
def api_get_order(order_id):
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    return jsonify(dict_from_row(row))


@app.route("/api/orders/<order_id>", methods=["PUT"])
def api_update_order(order_id):
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "error": "Order not found"}), 404
    if row["status"] not in ("registered", "failed", "cancelled"):
        return jsonify({"success": False, "error": "Only registered/failed/cancelled orders can be edited"}), 400

    data = request.json
    now = datetime.now().isoformat()

    order_type = data.get("order_type", row["order_type"])
    if order_type == "longitudinal" and not longitudinal_feature_enabled():
        if (row["order_type"] or "") != "longitudinal":
            return jsonify({"success": False, "error": "Longitudinal 분석이 설정에서 비활성화되어 있습니다."}), 403
    reuse_wid = (data.get("reuse_work_order_id") or "").strip() if "reuse_work_order_id" in data else None
    if reuse_wid is not None and reuse_wid:
        if order_type != "longitudinal":
            return jsonify({"success": False, "error": "reuse_work_order_id는 Longitudinal에서만 사용할 수 있습니다."}), 400
        src = db.execute("SELECT * FROM orders WHERE id=?", (reuse_wid,)).fetchone()
        if not src:
            return jsonify({"success": False, "error": "reuse_work_order_id: 오더를 찾을 수 없습니다."}), 400
        src = dict_from_row(src)
        if src.get("status") != "completed":
            return jsonify({"success": False, "error": "재사용할 오더는 분석 완료(completed) 상태여야 합니다."}), 400
        data["sample_name"] = src["sample_name"]
        data["r1_fastq"] = src["r1_fastq"]
        data["r2_fastq"] = src["r2_fastq"]

    editable = [
        "order_name", "patient_name", "patient_dob", "chart_number",
        "department", "doctor_name", "diagnosis", "doctor_comment",
        "sample_name", "r1_fastq", "r2_fastq", "reference", "profile",
        "bed_file", "delete_intermediate",
        "order_type", "baseline_order_id", "germline_order_id", "followup_order_ids",
        "reuse_work_order_id",
        "use_umi",
    ]
    sets = ["updated_at=?"]
    params = [now]
    for col in editable:
        if col in data:
            val = data[col]
            if col == "reference":
                val = val if val in ("hg38", "hg19") else "hg38"
            if col == "followup_order_ids" and isinstance(val, list):
                val = ",".join(val)
            if col == "use_umi":
                val = (val or "").strip().upper()
                if val not in ("Y", "N", ""):
                    val = ""
            sets.append(f"{col}=?")
            params.append(val)

    if "af_threshold" in data:
        sets.append("af_threshold=?")
        params.append(float(data["af_threshold"]))

    params.append(order_id)
    db.execute(f"UPDATE orders SET {', '.join(sets)} WHERE id=?", params)
    db.commit()
    return jsonify({"success": True})


@app.route("/api/orders/<order_id>/start", methods=["POST"])
def api_start_order(order_id):
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "error": "Order not found"}), 404
    order = dict_from_row(row)
    try:
        cid = start_analysis(
            order, force=False, resume=True, started_by_user_id=session.get("user_id") or "",
        )
        return jsonify({"success": True, "container_id": cid})
    except Exception as e:
        now = datetime.now().isoformat()
        db.execute("UPDATE orders SET status='failed', error_message=?, updated_at=? WHERE id=?",
                   (str(e), now, order_id))
        db.commit()
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/orders/<order_id>/stop", methods=["POST"])
def api_stop_order(order_id):
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "error": "Order not found"}), 404
    now = datetime.now().isoformat()
    container_name = (row["container_name"] or "").strip() or f"nxt_{row['sample_name']}_{order_id[:14]}"
    subprocess.run(["docker", "stop", container_name], capture_output=True, timeout=30)
    _archive_and_remove_container(order_id, row["sample_name"], container_name)
    _cleanup_nf_lock(order_id)
    db.execute("UPDATE orders SET status='cancelled', updated_at=? WHERE id=?", (now, order_id))
    db.commit()
    return jsonify({"success": True})


@app.route("/api/orders/<order_id>/rerun", methods=["POST"])
def api_rerun_order(order_id):
    """Rerun with Nextflow resume — leverages cached results."""
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "error": "Order not found"}), 404
    order = dict_from_row(row)
    container_name = (order.get("container_name") or "").strip() or f"nxt_{order['sample_name']}_{order_id[:14]}"
    subprocess.run(["docker", "rm", "-f", container_name], capture_output=True, timeout=10)
    _cleanup_nf_lock(order_id)
    try:
        cid = start_analysis(
            order, force=False, resume=True, started_by_user_id=session.get("user_id") or "",
        )
        return jsonify({"success": True, "container_id": cid})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/orders/<order_id>/force", methods=["POST"])
def api_force_order(order_id):
    """Force run from the beginning — wipe work dir and results."""
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "error": "Order not found"}), 404
    order = dict_from_row(row)
    if (order.get("reuse_work_order_id") or "").strip():
        return jsonify({
            "success": False,
            "error": "완료 오더 work 재사용(Longitudinal) 모드에서는 강제 재실행을 사용할 수 없습니다. 재실행(Rerun)으로 이어서 실행하세요.",
        }), 400
    container_name = (order.get("container_name") or "").strip() or f"nxt_{order['sample_name']}_{order_id[:14]}"
    subprocess.run(["docker", "rm", "-f", container_name], capture_output=True, timeout=10)
    _cleanup_nf_lock(order_id)
    if order.get("nf_work_dir") and os.path.isdir(order["nf_work_dir"]):
        shutil.rmtree(order["nf_work_dir"], ignore_errors=True)
    result_dir = os.path.join(RESULTS_DIR, order["sample_name"])
    if os.path.isdir(result_dir):
        shutil.rmtree(result_dir, ignore_errors=True)
    try:
        cid = start_analysis(
            order, force=True, resume=False, started_by_user_id=session.get("user_id") or "",
        )
        return jsonify({"success": True, "container_id": cid})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/orders/<order_id>/report_status", methods=["POST"])
def api_report_status(order_id):
    """External-integration endpoint: let an external CLI/program update an order's status.

    This allows a fully external Nextflow run (not managed by this Web UI's Docker control)
    to feed its lifecycle back into the order DB so that results are visible in the UI.

    Allowed transitions:
      registered → running   (analysis just started externally)
      running    → completed (analysis finished successfully)
      running    → failed    (analysis finished with error)
      registered → completed (shortcut: external run already done before this call)
      registered → failed

    Body (JSON):
      status        required  "running" | "completed" | "failed"
      error_message optional  error text when status="failed"
      nf_work_dir   optional  absolute host path to Nextflow work/ dir
      container_name optional container name for log lookup (if any)

    Returns { "success": true } on success.
    """
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "error": "Order not found"}), 404

    data = request.json or {}
    new_status = (data.get("status") or "").strip().lower()
    if new_status not in ("running", "completed", "failed"):
        return jsonify({"success": False, "error": "status must be 'running', 'completed', or 'failed'"}), 400

    current_status = row["status"]
    allowed_from = {
        "running":   ("registered",),
        "completed": ("registered", "running"),
        "failed":    ("registered", "running"),
    }
    if current_status not in allowed_from[new_status]:
        return jsonify({
            "success": False,
            "error": f"Cannot transition from '{current_status}' to '{new_status}'",
        }), 400

    now = datetime.now().isoformat()
    sets = ["status=?", "updated_at=?"]
    params = [new_status, now]

    if new_status == "running":
        sets.append("error_message=''")
    if new_status == "completed":
        sets.extend(["completed_at=?", "error_message=''"])
        params.append(now)
    if new_status == "failed":
        err = (data.get("error_message") or "").strip()
        sets.append("error_message=?")
        params.append(err)

    if "nf_work_dir" in data and data["nf_work_dir"]:
        sets.append("nf_work_dir=?")
        params.append(str(data["nf_work_dir"]))
    if "container_name" in data and data["container_name"]:
        sets.append("container_name=?")
        params.append(str(data["container_name"]))

    params.append(order_id)
    db.execute(f"UPDATE orders SET {', '.join(sets)} WHERE id=?", params)
    db.commit()
    return jsonify({"success": True, "order_id": order_id, "status": new_status})


@app.route("/api/orders/<order_id>/reannotate", methods=["POST"])
def api_reannotate(order_id):
    """Re-run SnpEff → SnpSift → VariantsToTable using the published *_vardict.vcf.
    Works regardless of delete_intermediate setting — only published VCF is needed.
    """
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "error": "Order not found"}), 404
    order = dict_from_row(row)

    if order.get("status") == "running":
        return jsonify({"success": False, "error": "분석이 진행 중입니다. 완료 후 시도하세요."}), 400

    sample = order["sample_name"]

    # Find published VarDict VCF
    vcf_candidates = glob.glob(
        os.path.join(RESULTS_DIR, sample, "**", f"{sample}_vardict.vcf"),
        recursive=True,
    )
    if not vcf_candidates:
        return jsonify({"success": False, "error": f"Published VarDict VCF를 찾을 수 없습니다: {sample}_vardict.vcf"}), 404

    vcf_host = vcf_candidates[0]
    # Convert host path to container path
    vcf_container = vcf_host.replace(RESULTS_DIR, "/work_nxt/results", 1)

    try:
        start_analysis(
            order,
            force=False,
            resume=False,
            started_by_user_id=session.get("user_id") or "",
            extra_nf_params=["--reannotate_vcf", vcf_container],
        )
        return jsonify({"success": True, "message": "Reannotation을 시작했습니다. SnpEff → SnpSift → VariantsToTable만 재실행합니다."})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/orders/<order_id>/promote_to_longitudinal", methods=["POST"])
def api_promote_to_longitudinal(order_id):
    """Convert an already-completed Followup order in-place to a Longitudinal
    order and trigger a Nextflow `-resume` run.

    Use case: Baseline, Germline, AND a Followup order have all been analyzed
    as regular orders. The user later decides to add Longitudinal comparison
    for that Followup. Instead of creating a new order, the existing Followup
    order is "promoted":

      - order_type           = 'longitudinal'
      - baseline_order_id    = <chosen completed baseline>
      - germline_order_id    = <chosen completed germline>
      - followup_order_ids   = (optional) prior followups for VAF trend
      - reuse_work_order_id  = '' (use this order's own work dir)

    Because work-dir / NXF_HOME / samplesheet are unchanged, Nextflow's
    -resume hits the cache and only the new processes (select_reporter,
    longitudinal) actually execute.
    """
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "error": "Order not found"}), 404
    order = dict_from_row(row)

    if (order.get("status") or "") != "completed":
        return jsonify({
            "success": False,
            "error": "분석 완료(completed) 상태의 오더만 Longitudinal로 변환할 수 있습니다.",
        }), 400
    if (order.get("order_type") or "") == "longitudinal":
        return jsonify({"success": False, "error": "이미 Longitudinal 오더입니다."}), 400
    if not longitudinal_feature_enabled():
        return jsonify({"success": False, "error": "Longitudinal 분석이 설정에서 비활성화되어 있습니다."}), 403

    data = request.json or {}
    baseline_id = (data.get("baseline_order_id") or "").strip()
    germline_id = (data.get("germline_order_id") or "").strip()
    if not baseline_id or not germline_id:
        return jsonify({
            "success": False,
            "error": "Baseline 오더와 Germline 오더를 모두 선택해야 합니다.",
        }), 400
    if baseline_id == order_id or germline_id == order_id:
        return jsonify({
            "success": False,
            "error": "Baseline/Germline 오더는 변환 대상 오더 자신일 수 없습니다.",
        }), 400

    for ref_id, label in ((baseline_id, "Baseline"), (germline_id, "Germline")):
        ref = db.execute("SELECT id, status FROM orders WHERE id=?", (ref_id,)).fetchone()
        if not ref:
            return jsonify({"success": False, "error": f"{label} 오더를 찾을 수 없습니다: {ref_id}"}), 400
        if ref["status"] != "completed":
            return jsonify({
                "success": False,
                "error": f"{label} 오더는 분석 완료(completed) 상태여야 합니다.",
            }), 400

    followup_ids_raw = data.get("followup_order_ids", "")
    if isinstance(followup_ids_raw, list):
        followup_ids = ",".join([fid.strip() for fid in followup_ids_raw if fid and str(fid).strip()])
    else:
        followup_ids = str(followup_ids_raw or "").strip()
    # Reject self-reference in followup list
    if followup_ids:
        if order_id in [fid.strip() for fid in followup_ids.split(",")]:
            return jsonify({
                "success": False,
                "error": "이전 Followup 목록에 변환 대상 오더 자신을 포함할 수 없습니다.",
            }), 400

    now = datetime.now().isoformat()
    db.execute(
        """UPDATE orders
              SET order_type='longitudinal',
                  baseline_order_id=?,
                  germline_order_id=?,
                  followup_order_ids=?,
                  reuse_work_order_id='',
                  status='queued',
                  error_message='',
                  updated_at=?
            WHERE id=?""",
        (baseline_id, germline_id, followup_ids, now, order_id),
    )
    db.commit()

    # Re-fetch with updated columns and trigger a resume run.
    order = dict_from_row(db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone())
    container_name = (order.get("container_name") or "").strip() or f"nxt_{order['sample_name']}_{order_id[:14]}"
    subprocess.run(["docker", "rm", "-f", container_name], capture_output=True, timeout=10)
    _cleanup_nf_lock(order_id)

    # ── Bypass mode: use published BAM + annotated TXT when work dir is sparse ──
    # When delete_intermediate=Y cleaned the work dir, we can skip the heavy
    # FASTQ→BAM→VarDict steps by pointing Nextflow directly at the published
    # outputs that were preserved in results/<sample>/output/.
    sample = order["sample_name"]
    bypass_params = []

    bam_pattern  = os.path.join(RESULTS_DIR, sample, "output", "bam", f"{sample}_clipped_sorted.bam")
    ann_pattern  = os.path.join(RESULTS_DIR, sample, "output", f"{sample}_vardict_annotated_vcf.txt")

    # Also allow the annotated txt that may live under a subdirectory (output_subdir)
    import glob as _glob
    if not os.path.isfile(ann_pattern):
        candidates = _glob.glob(os.path.join(RESULTS_DIR, sample, "output", "**", f"{sample}_vardict_annotated_vcf.txt"), recursive=True)
        if candidates:
            ann_pattern = candidates[0]

    if os.path.isfile(bam_pattern) and os.path.isfile(ann_pattern):
        bai_path = bam_pattern + ".bai"
        if not os.path.isfile(bai_path):
            bai_candidates = _glob.glob(os.path.join(RESULTS_DIR, sample, "output", "bam", "*.bai"))
            bai_path = bai_candidates[0] if bai_candidates else bam_pattern + ".bai"

        # Container-internal paths (/work_nxt maps to host BASE_DIR)
        results_rel = os.path.relpath(RESULTS_DIR, BASE_DIR)
        container_bam = f"/work_nxt/{results_rel}/{sample}/output/bam/{os.path.basename(bam_pattern)}"
        container_bai = f"/work_nxt/{results_rel}/{sample}/output/bam/{os.path.basename(bai_path)}"
        container_ann = f"/work_nxt/{os.path.relpath(ann_pattern, BASE_DIR)}"

        bypass_params = [
            "--precomputed_bam",     container_bam,
            "--precomputed_bai",     container_bai,
            "--precomputed_ann_txt", container_ann,
        ]
        bypass_mode = True
    else:
        bypass_mode = False

    # Warn if the work directory has few cached tasks (delete_intermediate may have cleaned it up)
    # and we also could not find published outputs for bypass mode.
    work_dir = os.path.join(BASE_DIR, "work", order_id)
    cached_count = 0
    if os.path.isdir(work_dir):
        try:
            cached_count = len(_glob.glob(os.path.join(work_dir, "**", ".exitcode"), recursive=True))
        except Exception:
            pass

    if bypass_mode:
        resume_warning = ""  # bypass: no full re-run needed
    elif cached_count < 10:
        resume_warning = (
            "중간 파일이 삭제된 상태여서 전체 파이프라인을 처음부터 재실행합니다 "
            "(delete_intermediate 옵션이 켜져 있었던 것으로 추정됩니다). "
            "완료 후 Longitudinal 단계가 자동으로 이어집니다."
        )
    else:
        resume_warning = ""

    try:
        cid = start_analysis(
            order, force=False, resume=(not bypass_mode), started_by_user_id=session.get("user_id") or "",
            extra_nf_params=bypass_params if bypass_mode else None,
        )
        mode_msg = "Bypass 모드 (published BAM 재사용): SELECT_REPORTER + Longitudinal 단계만 실행합니다." if bypass_mode else "Nextflow -resume 으로 재실행합니다."
        return jsonify({
            "success": True,
            "container_id": cid,
            "message": f"Longitudinal 분석을 시작했습니다. {mode_msg}",
            "warning": resume_warning,
            "bypass_mode": bypass_mode,
        })
    except Exception as e:
        db.execute(
            "UPDATE orders SET status='failed', error_message=?, updated_at=? WHERE id=?",
            (str(e), datetime.now().isoformat(), order_id),
        )
        db.commit()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/orders/<order_id>/delete", methods=["DELETE"])
def api_delete_order(order_id):
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "error": "Order not found"}), 404
    order = dict_from_row(row)
    if order["status"] == "running":
        container_name = (order.get("container_name") or "").strip() or f"nxt_{order['sample_name']}_{order_id[:14]}"
        subprocess.run(["docker", "rm", "-f", container_name], capture_output=True, timeout=10)
        _cleanup_nf_lock(order_id)
    purge_order_disk_assets(order)
    db.execute("DELETE FROM orders WHERE id=?", (order_id,))
    db.commit()
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Routes — Logs
# ---------------------------------------------------------------------------
_ANSI_RE = re.compile(r"\x1b\[[0-9;]*[A-Za-z]|\r")


def _clean_nf_log(raw: str) -> str:
    """Strip ANSI escapes and collapse repeated Nextflow progress blocks."""
    text = _ANSI_RE.sub("", raw)
    lines = text.split("\n")

    messages = []
    progress = {}
    last_executor_line = ""

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith("executor >"):
            last_executor_line = stripped
            continue
        if stripped.startswith("[") and ("|" in stripped) and ("of" in stripped or "cached" in stripped):
            key = stripped.split("]")[0] + "]"
            progress[key] = stripped
            continue
        if stripped.startswith("Plus ") and "more processes waiting" in stripped:
            progress["__plus__"] = stripped
            continue
        messages.append(line)

    result_lines = []
    for line in messages:
        result_lines.append(line)

    if progress:
        result_lines.append("")
        if last_executor_line:
            result_lines.append(last_executor_line)
        for key in sorted(progress.keys()):
            if key != "__plus__":
                result_lines.append(progress[key])
        if "__plus__" in progress:
            result_lines.append(progress["__plus__"])

    return "\n".join(result_lines)


@app.route("/api/orders/<order_id>/logs")
def api_order_logs(order_id):
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "error": "Order not found"}), 404

    order_snapshot = format_order_snapshot(row)
    created_by = (row["created_by_user_id"] or "").strip() or "-"
    analysis_by = (row["analysis_by_user_id"] or "").strip() or "-"

    sample = row["sample_name"]
    tail = int(request.args.get("tail", "500"))

    # Use stored container_name; fall back to computed name
    container_name = (row["container_name"] or "").strip() \
        or f"nxt_{sample}_{order_id[:14]}"
    raw = ""
    try:
        r = subprocess.run(
            ["docker", "logs", "--tail", str(tail), container_name],
            capture_output=True, text=True, timeout=10,
        )
        if r.returncode == 0 and (r.stdout or r.stderr):
            raw = r.stdout + r.stderr
    except Exception:
        pass

    if not raw:
        log_candidates = [
            os.path.join(LOG_DIR, f"{sample}_{order_id}_nf.log"),
            os.path.join(LOG_DIR, f"{sample}.log"),
        ]
        for lf in log_candidates:
            if os.path.isfile(lf):
                try:
                    with open(lf, "r", errors="replace") as f:
                        raw = f.read()
                except Exception:
                    pass
                break

    log_text = "No logs available yet."
    if raw:
        log_text = _clean_nf_log(raw)

    return jsonify({
        "success": True,
        "order_snapshot": order_snapshot,
        "created_by_user_id": created_by,
        "analysis_by_user_id": analysis_by,
        "logs": log_text,
    })


# ---------------------------------------------------------------------------
# Routes — Files
# ---------------------------------------------------------------------------
@app.route("/api/fastq_files")
def api_fastq_files():
    return jsonify(browse_fastq(request.args.get("path", "")))


@app.route("/api/bed_files")
def api_bed_files():
    return jsonify(browse_bed(request.args.get("path", "")))


@app.route("/api/results_image/<path:rel_path>")
def api_results_image(rel_path):
    """Serve result image files (e.g. expression plots) by relative path under RESULTS_DIR."""
    _require_auth()
    # Prevent directory traversal
    safe_path = os.path.normpath(os.path.join(RESULTS_DIR, rel_path))
    if not safe_path.startswith(os.path.normpath(RESULTS_DIR)):
        return jsonify({"error": "forbidden"}), 403
    if not os.path.isfile(safe_path):
        return jsonify({"error": "not found"}), 404
    mime = "image/png"
    if safe_path.endswith(".jpg") or safe_path.endswith(".jpeg"):
        mime = "image/jpeg"
    elif safe_path.endswith(".svg"):
        mime = "image/svg+xml"
    elif safe_path.endswith(".html"):
        mime = "text/html"
    return send_file(safe_path, mimetype=mime)


@app.route("/api/download/<sample_name>/<file_type>")
def api_download(sample_name, file_type):
    try:
        if file_type == "vcf":
            fp = glob.glob(os.path.join(RESULTS_DIR, sample_name, "**", f"{sample_name}*.vcf"), recursive=True)
            fp = fp[0] if fp else ""
        elif file_type == "txt":
            fp = glob.glob(os.path.join(RESULTS_DIR, sample_name, "**", f"{sample_name}*annotated*.txt"), recursive=True)
            fp = fp[0] if fp else ""
        elif file_type == "log":
            candidates = glob.glob(os.path.join(LOG_DIR, f"{sample_name}_*_nf.log"))
            fp = candidates[0] if candidates else ""
        # ── RNA-seq specific download types ──────────────────────────────────
        elif file_type == "counts":
            # featureCounts raw count matrix
            candidates = glob.glob(os.path.join(RESULTS_DIR, sample_name, "featureCounts", f"{sample_name}_counts.txt"))
            fp = candidates[0] if candidates else ""
        elif file_type == "expression":
            # CPM/TPM expression summary
            candidates = glob.glob(os.path.join(RESULTS_DIR, sample_name, "expression_plots", f"{sample_name}_expression_summary.tsv"))
            fp = candidates[0] if candidates else ""
        elif file_type == "multiqc":
            # MultiQC HTML report
            candidates = glob.glob(os.path.join(RESULTS_DIR, "MultiQC", "multiqc_report.html"))
            fp = candidates[0] if candidates else ""
        elif file_type == "rna_qc_json":
            # RNA QC JSON
            candidates = glob.glob(os.path.join(RESULTS_DIR, sample_name, "QC_report", f"{sample_name}_rna_qc.json"))
            fp = candidates[0] if candidates else ""
        elif file_type == "fusion":
            # STAR-Fusion results (prefer coding_effect abridged)
            fusion_dir = os.path.join(RESULTS_DIR, sample_name, "star_fusion")
            fp = ""
            for pat in [
                f"{sample_name}_star-fusion.fusion_predictions.abridged.coding_effect.tsv",
                f"{sample_name}_star-fusion.fusion_predictions.abridged.tsv",
            ]:
                candidate = os.path.join(fusion_dir, pat)
                if os.path.isfile(candidate):
                    fp = candidate
                    break
        elif file_type == "expression_xlsx":
            # Expression summary TSV → Excel (expressed genes only, count > 0)
            import io
            tsv_candidates = glob.glob(os.path.join(
                RESULTS_DIR, sample_name, "expression_plots",
                f"{sample_name}_expression_summary.tsv"))
            if not tsv_candidates or not os.path.isfile(tsv_candidates[0]):
                return jsonify({"error": "Expression file not found"}), 404
            import pandas as pd
            df = pd.read_csv(tsv_candidates[0], sep="\t")
            # Only include expressed genes (count > 0) — reduces ~62k rows to ~15-20k
            df = df[df["count"] > 0].copy()
            df.sort_values("TPM", ascending=False, inplace=True)
            col_map = {"gene_id": "Gene ID", "gene_symbol": "Gene Symbol",
                       "length": "Length (bp)", "count": "Raw Count",
                       "CPM": "CPM", "TPM": "TPM", "log2CPM": "log2(CPM+1)"}
            df.rename(columns={k: v for k, v in col_map.items() if k in df.columns}, inplace=True)
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Expression")
                ws = writer.sheets["Expression"]
                for col in ws.columns:
                    max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
            buf.seek(0)
            resp = make_response(buf.read())
            resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            resp.headers["Content-Disposition"] = f'attachment; filename="{sample_name}_expression.xlsx"'
            return resp
        else:
            return jsonify({"error": "Invalid file type"}), 400
        if fp and os.path.isfile(fp):
            return send_file(fp, as_attachment=True)
        return jsonify({"error": "File not found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/orders/<order_id>/qc_data")
def api_qc_data(order_id):
    """Collect all QC metrics for the given order."""
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"error": "Order not found"}), 404
    order = dict_from_row(row)
    sample = order["sample_name"]

    qc_dir = os.path.join(RESULTS_DIR, sample, "QC_report")
    trim_dir = os.path.join(RESULTS_DIR, sample, "trimming")

    def _parse_picard(filepath):
        """Parse a Picard-style metrics file: return list of dicts (one per data row)."""
        if not os.path.isfile(filepath):
            return None
        rows = []
        header = None
        in_metrics = False
        with open(filepath, "r", errors="replace") as fh:
            for line in fh:
                line = line.rstrip("\n")
                if line.startswith("## METRICS CLASS"):
                    in_metrics = True
                    continue
                if line.startswith("## HISTOGRAM"):
                    break
                if in_metrics:
                    if header is None:
                        header = line.split("\t")
                    elif line.strip():
                        vals = line.split("\t")
                        rows.append(dict(zip(header, vals)))
                    else:
                        break
        return rows if rows else None

    def _read_count(filepath):
        if not os.path.isfile(filepath):
            return None
        with open(filepath, "r") as fh:
            text = fh.read().strip()
        for line in text.split("\n"):
            line = line.strip()
            if not line:
                continue
            if line.startswith("Tool returned:"):
                val = line.split(":", 1)[1].strip()
                if val:
                    try:
                        return int(val)
                    except ValueError:
                        continue
            else:
                try:
                    return int(line)
                except ValueError:
                    continue
        return None

    result = {"success": True, "sample_name": sample, "order_name": order.get("order_name", "")}

    # Subsampling info (stored at analysis-start time)
    sub_info_raw = order.get("subsample_info", "") or ""
    if sub_info_raw:
        try:
            result["subsample_info"] = json.loads(sub_info_raw)
        except Exception:
            pass

    # 1. fastp
    fastp_path = os.path.join(trim_dir, "fastp.json")
    if os.path.isfile(fastp_path):
        try:
            with open(fastp_path, "r") as fh:
                fj = json.load(fh)
            s_before = fj.get("summary", {}).get("before_filtering", {})
            s_after = fj.get("summary", {}).get("after_filtering", {})
            result["fastp"] = {
                "before_total_reads": s_before.get("total_reads", 0),
                "before_total_bases": s_before.get("total_bases", 0),
                "before_q20_rate": s_before.get("q20_rate", 0),
                "before_q30_rate": s_before.get("q30_rate", 0),
                "before_gc_content": s_before.get("gc_content", 0),
                "after_total_reads": s_after.get("total_reads", 0),
                "after_total_bases": s_after.get("total_bases", 0),
                "after_q20_rate": s_after.get("q20_rate", 0),
                "after_q30_rate": s_after.get("q30_rate", 0),
                "after_gc_content": s_after.get("gc_content", 0),
                "after_read1_mean_length": s_after.get("read1_mean_length", 0),
                "after_read2_mean_length": s_after.get("read2_mean_length", 0),
            }
            fc = fj.get("filtering_result", {})
            if fc:
                result["fastp"]["passed_filter_reads"] = fc.get("passed_filter_reads", 0)
                result["fastp"]["low_quality_reads"] = fc.get("low_quality_reads", 0)
                result["fastp"]["too_short_reads"] = fc.get("too_short_reads", 0)
                result["fastp"]["adapter_trimmed_reads"] = fj.get("adapter_cutting", {}).get("adapter_trimmed_reads", 0)
        except Exception:
            pass

    # 2. Alignment summary (umi_deduped and aligned)
    for label in ["umi_deduped", "aligned"]:
        fp = os.path.join(qc_dir, f"{sample}_alignment_metrics_{label}.txt")
        rows = _parse_picard(fp)
        if rows:
            pair_row = next((r for r in rows if r.get("CATEGORY") == "PAIR"), rows[-1])
            result[f"alignment_{label}"] = {
                "total_reads": pair_row.get("TOTAL_READS", ""),
                "pf_reads_aligned": pair_row.get("PF_READS_ALIGNED", ""),
                "pct_pf_reads_aligned": pair_row.get("PCT_PF_READS_ALIGNED", ""),
                "pf_mismatch_rate": pair_row.get("PF_MISMATCH_RATE", ""),
                "pct_chimeras": pair_row.get("PCT_CHIMERAS", ""),
                "pct_adapter": pair_row.get("PCT_ADAPTER", ""),
                "mean_read_length": pair_row.get("MEAN_READ_LENGTH", ""),
                "pct_reads_aligned_in_pairs": pair_row.get("PCT_READS_ALIGNED_IN_PAIRS", ""),
                "strand_balance": pair_row.get("STRAND_BALANCE", ""),
            }

    # 3. Insert size
    for label in ["umi_deduped", "aligned"]:
        fp = os.path.join(qc_dir, f"{sample}_insert_size_metrics_{label}.txt")
        rows = _parse_picard(fp)
        if rows:
            r0 = rows[0]
            result[f"insert_size_{label}"] = {
                "median_insert_size": r0.get("MEDIAN_INSERT_SIZE", ""),
                "mean_insert_size": r0.get("MEAN_INSERT_SIZE", ""),
                "standard_deviation": r0.get("STANDARD_DEVIATION", ""),
                "min_insert_size": r0.get("MIN_INSERT_SIZE", ""),
                "max_insert_size": r0.get("MAX_INSERT_SIZE", ""),
                "mode_insert_size": r0.get("MODE_INSERT_SIZE", ""),
                "read_pairs": r0.get("READ_PAIRS", ""),
            }

    # 4. MarkDuplicates
    fp = os.path.join(qc_dir, f"{sample}_markduplicates_metrics_gatk.txt")
    rows = _parse_picard(fp)
    if rows:
        r0 = rows[0]
        result["duplicates"] = {
            "read_pairs_examined": r0.get("READ_PAIRS_EXAMINED", ""),
            "read_pair_duplicates": r0.get("READ_PAIR_DUPLICATES", ""),
            "read_pair_optical_duplicates": r0.get("READ_PAIR_OPTICAL_DUPLICATES", ""),
            "percent_duplication": r0.get("PERCENT_DUPLICATION", ""),
            "estimated_library_size": r0.get("ESTIMATED_LIBRARY_SIZE", ""),
        }

    # 5. On-target reads
    for label in ["umi_deduped", "aligned"]:
        fp = os.path.join(qc_dir, f"{sample}_ontarget_reads_{label}.txt")
        count = _read_count(fp)
        if count is not None:
            result[f"ontarget_{label}"] = count

    # 6. HS metrics
    for label in ["umi_deduped", "aligned"]:
        fp = os.path.join(qc_dir, f"{sample}_hs_metrics_{label}.txt")
        rows = _parse_picard(fp)
        if rows:
            r0 = rows[0]
            result[f"hs_metrics_{label}"] = {
                "mean_target_coverage": r0.get("MEAN_TARGET_COVERAGE", ""),
                "median_target_coverage": r0.get("MEDIAN_TARGET_COVERAGE", ""),
                "max_target_coverage": r0.get("MAX_TARGET_COVERAGE", ""),
                "pct_selected_bases": r0.get("PCT_SELECTED_BASES", ""),
                "fold_enrichment": r0.get("FOLD_ENRICHMENT", ""),
                "zero_cvg_targets_pct": r0.get("ZERO_CVG_TARGETS_PCT", ""),
                "fold_80_base_penalty": r0.get("FOLD_80_BASE_PENALTY", ""),
                "pct_target_bases_1x": r0.get("PCT_TARGET_BASES_1X", ""),
                "pct_target_bases_10x": r0.get("PCT_TARGET_BASES_10X", ""),
                "pct_target_bases_20x": r0.get("PCT_TARGET_BASES_20X", ""),
                "pct_target_bases_30x": r0.get("PCT_TARGET_BASES_30X", ""),
                "pct_target_bases_50x": r0.get("PCT_TARGET_BASES_50X", ""),
                "pct_target_bases_100x": r0.get("PCT_TARGET_BASES_100X", ""),
                "pct_target_bases_250x": r0.get("PCT_TARGET_BASES_250X", ""),
                "pct_target_bases_500x": r0.get("PCT_TARGET_BASES_500X", ""),
                "pct_target_bases_1000x": r0.get("PCT_TARGET_BASES_1000X", ""),
                "total_reads": r0.get("TOTAL_READS", ""),
                "pf_unique_reads": r0.get("PF_UNIQUE_READS", ""),
                "on_target_bases": r0.get("ON_TARGET_BASES", ""),
                "pf_bases_aligned": r0.get("PF_BASES_ALIGNED", ""),
                "target_territory": r0.get("TARGET_TERRITORY", ""),
                "bait_territory": r0.get("BAIT_TERRITORY", ""),
                "on_bait_bases": r0.get("ON_BAIT_BASES", ""),
                "pct_usable_bases_on_target": r0.get("PCT_USABLE_BASES_ON_TARGET", ""),
            }

    # 7. Mismatch rate
    fp = os.path.join(qc_dir, f"{sample}_mismatch_rate.csv")
    if os.path.isfile(fp):
        try:
            with open(fp, "r") as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    result["mismatch_rate"] = {k: v for k, v in row.items()}
                    break
        except Exception:
            pass

    # ── 8. RNA-seq QC (parsed from _rna_qc.json generated by RNASEQ_QC_SUMMARY) ──
    rna_qc_path = os.path.join(qc_dir, f"{sample}_rna_qc.json")
    if os.path.isfile(rna_qc_path):
        try:
            with open(rna_qc_path, "r") as fh:
                rna_qc = json.load(fh)
            result["rna_qc"] = rna_qc
            # Flatten STAR summary for easy access in UI
            star = rna_qc.get("star", {})
            result["rna_star"] = {
                "total_reads":            star.get("total_reads", ""),
                "uniquely_mapped":         star.get("uniquely_mapped", ""),
                "uniquely_mapped_pct":     star.get("uniquely_mapped_pct", ""),
                "multi_mapped":            star.get("multi_mapped", ""),
                "multi_mapped_pct":        star.get("multi_mapped_pct", ""),
                "unmapped_too_short_pct":  star.get("unmapped_too_short_pct", ""),
                "avg_mapped_length":       star.get("avg_mapped_length", ""),
                "mismatch_rate_per_base":  star.get("mismatch_rate_per_base", ""),
                "splices_total":           star.get("splices_total", ""),
                "splices_annotated":       star.get("splices_annotated", ""),
            }
            # featureCounts assignment summary
            fc = rna_qc.get("featurecounts", {})
            result["rna_featurecounts"] = fc
            # flagstat
            result["rna_flagstat"] = rna_qc.get("flagstat", {})
            # fastp (override/supplement existing fastp block if present)
            if not result.get("fastp") and rna_qc.get("fastp"):
                result["fastp"] = rna_qc["fastp"]
        except Exception:
            pass

    # ── 9. RNA expression summary (top genes, CPM/TPM table) ─────────────────
    expr_dir = os.path.join(RESULTS_DIR, sample, "expression_plots")
    expr_summary_path = os.path.join(expr_dir, f"{sample}_expression_summary.tsv")
    if os.path.isfile(expr_summary_path):
        try:
            import csv as _csv
            with open(expr_summary_path, "r") as fh:
                reader = _csv.DictReader(fh, delimiter="\t")
                rows = []
                for i, row in enumerate(reader):
                    if i >= 200:  # Return top 200 genes for UI display
                        break
                    rows.append(dict(row))
            result["rna_expression"] = {
                "top_genes": rows,
                "total_genes": sum(1 for _ in open(expr_summary_path)) - 1,
            }
        except Exception:
            pass

    # ── 10. RNA expression plot images (relative paths for UI) ───────────────
    if os.path.isdir(expr_dir):
        plot_files = glob.glob(os.path.join(expr_dir, f"{sample}_*.png"))
        result["rna_plots"] = [
            os.path.relpath(p, RESULTS_DIR) for p in sorted(plot_files)
        ]
        interactive_files = glob.glob(os.path.join(expr_dir, f"{sample}_interactive_*.html"))
        result["rna_interactive_plots"] = [
            os.path.relpath(p, RESULTS_DIR) for p in sorted(interactive_files)
        ]

    # ── 11. STAR-Fusion results ───────────────────────────────────────────────
    fusion_dir = os.path.join(RESULTS_DIR, sample, "star_fusion")
    # Prefer coding_effect abridged file; fall back to plain abridged
    for fusion_pattern in [
        f"{sample}_star-fusion.fusion_predictions.abridged.coding_effect.tsv",
        f"{sample}_star-fusion.fusion_predictions.abridged.tsv",
    ]:
        fusion_path = os.path.join(fusion_dir, fusion_pattern)
        if os.path.isfile(fusion_path):
            try:
                fusions = []
                with open(fusion_path) as fh:
                    header = None
                    for line in fh:
                        line = line.rstrip("\n")
                        if line.startswith("#"):
                            header = line.lstrip("#").split("\t")
                            continue
                        if header is None:
                            continue
                        parts = line.split("\t")
                        row = dict(zip(header, parts))
                        fusions.append(row)
                result["rna_fusion"] = {
                    "fusions": fusions,
                    "count": len(fusions),
                    "source_file": fusion_pattern,
                }
            except Exception:
                pass
            break

    # Check which QC files exist
    result["has_qc"] = os.path.isdir(qc_dir) and bool(os.listdir(qc_dir))
    result["panel_type"] = order.get("panel_type", "exome")

    # ── Computed QC summary (6 key metrics) ──────────────────────────────────
    try:
        fp_s   = result.get("fastp", {})
        al_s   = result.get("alignment_aligned", {})
        dup_s  = result.get("duplicates", {})
        hs_s   = result.get("hs_metrics_umi_deduped") or result.get("hs_metrics_aligned") or {}
        ot_al  = result.get("ontarget_aligned")

        before_bases = float(fp_s.get("before_total_bases") or 0)
        before_reads = float(fp_s.get("before_total_reads") or 0)
        pf_aligned   = float(al_s.get("pf_reads_aligned") or 0)
        total_reads  = float(al_s.get("total_reads") or 0)
        dup_pct      = float(dup_s.get("percent_duplication") or 0)
        mean_cov     = float(hs_s.get("mean_target_coverage") or 0)
        pct_100x     = float(hs_s.get("pct_target_bases_100x") or 0)

        summary = {}
        if before_bases > 0:
            summary["throughput_mb"] = round(before_bases / 1e6, 1)
        if pf_aligned > 0 and before_reads > 0:
            # True mapping rate: aligned reads vs original raw reads
            summary["mapped_pct"] = round(pf_aligned / before_reads * 100, 2)
        if ot_al is not None and total_reads > 0:
            summary["ontarget_pct"] = round(ot_al / total_reads * 100, 2)
        if mean_cov > 0:
            summary["ontarget_coverage_x"] = round(mean_cov, 1)
        if pct_100x > 0:
            summary["uniformity_100x"] = round(pct_100x * 100, 2)
        if dup_pct >= 0 and dup_s:
            summary["duplicated_pct"] = round(dup_pct * 100, 2)

        if summary:
            result["qc_summary"] = summary
    except Exception:
        pass

    return jsonify(result)


@app.route("/api/orders/<order_id>/qc_report.txt")
def api_qc_report_txt(order_id):
    """Generate a plain-text QC summary report for download."""
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return ("Order not found", 404)
    order = dict_from_row(row)

    # Reuse api_qc_data logic by calling it internally
    from flask import current_app
    with current_app.test_request_context(f"/api/orders/{order_id}/qc_data"):
        resp = api_qc_data(order_id)
        if hasattr(resp, "get_json"):
            d = resp.get_json()
        else:
            import json as _json
            d = _json.loads(resp.data)

    lines = []
    def h(title):
        lines.append("")
        lines.append("=" * 60)
        lines.append(f"  {title}")
        lines.append("=" * 60)
    def row_kv(k, v):
        lines.append(f"  {k:<40} {v}")

    lines.append("Roche_nxt QC Summary Report")
    lines.append(f"Sample  : {d.get('sample_name','')}")
    lines.append(f"Order   : {d.get('order_name','')}")
    lines.append(f"Order ID: {order_id}")
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # QC Key Metrics Summary
    qs = d.get("qc_summary")
    if qs:
        h("QC Key Metrics")
        if "throughput_mb"      in qs: row_kv("Throughput (Mb)",               f"{qs['throughput_mb']:,.1f}")
        if "mapped_pct"         in qs: row_kv("Alignment Rate (%)",             f"{qs['mapped_pct']:.2f}%")
        if "ontarget_pct"       in qs: row_kv("On-Target (%)",                  f"{qs['ontarget_pct']:.2f}%")
        if "ontarget_coverage_x" in qs: row_kv("On-Target Coverage (x)",        f"{qs['ontarget_coverage_x']:.1f}x")
        if "uniformity_100x"    in qs: row_kv("Coverage Uniformity (>=100x, %)", f"{qs['uniformity_100x']:.2f}%")
        if "duplicated_pct"     in qs: row_kv("Duplication (%)",                f"{qs['duplicated_pct']:.2f}%")

    # Subsampling
    sub = d.get("subsample_info")
    if sub and sub.get("enabled"):
        h("Subsampling")
        row_kv("R1 size (GB)", sub.get("r1_gb",""))
        row_kv("R2 size (GB)", sub.get("r2_gb",""))
        row_kv("Total size (GB)", sub.get("total_gb",""))
        row_kv("Threshold (GB)", sub.get("threshold_gb",""))
        row_kv("Target reads", sub.get("target_reads",""))
        row_kv("Seed", sub.get("seed",""))

    # fastp
    fp = d.get("fastp")
    if fp:
        h("Adapter Trimming (fastp)")
        row_kv("Reads before filtering", fp.get("before_total_reads",""))
        row_kv("Reads after filtering", fp.get("after_total_reads",""))
        row_kv("Passed filter reads", fp.get("passed_filter_reads",""))
        row_kv("Low quality reads (removed)", fp.get("low_quality_reads",""))
        row_kv("Too short reads (removed)", fp.get("too_short_reads",""))
        row_kv("Adapter trimmed reads", fp.get("adapter_trimmed_reads",""))
        row_kv("Q20 rate (before)", f"{float(fp.get('before_q20_rate',0))*100:.2f}%")
        row_kv("Q30 rate (before)", f"{float(fp.get('before_q30_rate',0))*100:.2f}%")
        row_kv("Q20 rate (after)", f"{float(fp.get('after_q20_rate',0))*100:.2f}%")
        row_kv("Q30 rate (after)", f"{float(fp.get('after_q30_rate',0))*100:.2f}%")
        row_kv("GC content (after)", f"{float(fp.get('after_gc_content',0))*100:.2f}%")
        row_kv("Mean read length R1 (after)", fp.get("after_read1_mean_length",""))
        row_kv("Mean read length R2 (after)", fp.get("after_read2_mean_length",""))

    # Alignment
    for label, title in [("aligned", "Alignment Metrics — Aligned (pre-dedup)"),
                          ("umi_deduped", "Alignment Metrics — UMI Deduped")]:
        al = d.get(f"alignment_{label}")
        if al:
            h(title)
            row_kv("Total reads", al.get("total_reads",""))
            row_kv("PF reads aligned", al.get("pf_reads_aligned",""))
            row_kv("% PF reads aligned", f"{float(al.get('pct_pf_reads_aligned',0))*100:.2f}%")
            row_kv("Mismatch rate", al.get("pf_mismatch_rate",""))
            row_kv("% chimeras", al.get("pct_chimeras",""))
            row_kv("% adapter", al.get("pct_adapter",""))
            row_kv("Mean read length", al.get("mean_read_length",""))
            row_kv("% reads aligned in pairs", f"{float(al.get('pct_reads_aligned_in_pairs',0))*100:.2f}%")
            row_kv("Strand balance", al.get("strand_balance",""))

    # Insert size
    for label, title in [("aligned", "Insert Size — Aligned"),
                          ("umi_deduped", "Insert Size — UMI Deduped")]:
        ins = d.get(f"insert_size_{label}")
        if ins:
            h(title)
            row_kv("Median insert size", ins.get("median_insert_size",""))
            row_kv("Mean insert size", ins.get("mean_insert_size",""))
            row_kv("Std deviation", ins.get("standard_deviation",""))
            row_kv("Min / Max", f"{ins.get('min_insert_size','')} / {ins.get('max_insert_size','')}")
            row_kv("Read pairs", ins.get("read_pairs",""))

    # Duplicates
    dup = d.get("duplicates")
    if dup:
        h("Duplicate Metrics (MarkDuplicates)")
        row_kv("Read pairs examined", dup.get("read_pairs_examined",""))
        row_kv("Read pair duplicates", dup.get("read_pair_duplicates",""))
        row_kv("Optical duplicates", dup.get("read_pair_optical_duplicates",""))
        row_kv("% duplication", f"{float(dup.get('percent_duplication',0))*100:.4f}%")
        row_kv("Estimated library size", dup.get("estimated_library_size",""))

    # On-target
    ot_al = d.get("ontarget_aligned")
    ot_umi = d.get("ontarget_umi_deduped")
    if ot_al is not None or ot_umi is not None:
        h("On-target Reads")
        if ot_al is not None:  row_kv("On-target reads (aligned)", f"{ot_al:,}")
        if ot_umi is not None: row_kv("On-target reads (UMI deduped)", f"{ot_umi:,}")

    # HS metrics
    for label, title in [("aligned", "Hybridization Selection Metrics — Aligned"),
                          ("umi_deduped", "Hybridization Selection Metrics — UMI Deduped")]:
        hs = d.get(f"hs_metrics_{label}")
        if hs:
            h(title)
            row_kv("Mean target coverage", hs.get("mean_target_coverage",""))
            row_kv("Median target coverage", hs.get("median_target_coverage",""))
            row_kv("Max target coverage", hs.get("max_target_coverage",""))
            row_kv("% selected bases", f"{float(hs.get('pct_selected_bases',0))*100:.2f}%")
            row_kv("Fold enrichment", hs.get("fold_enrichment",""))
            row_kv("Fold 80 base penalty", hs.get("fold_80_base_penalty",""))
            row_kv("% zero coverage targets", f"{float(hs.get('zero_cvg_targets_pct',0))*100:.2f}%")
            row_kv("% bases >=1x", f"{float(hs.get('pct_target_bases_1x',0))*100:.2f}%")
            row_kv("% bases >=10x", f"{float(hs.get('pct_target_bases_10x',0))*100:.2f}%")
            row_kv("% bases >=20x", f"{float(hs.get('pct_target_bases_20x',0))*100:.2f}%")
            row_kv("% bases >=30x", f"{float(hs.get('pct_target_bases_30x',0))*100:.2f}%")
            row_kv("% bases >=50x", f"{float(hs.get('pct_target_bases_50x',0))*100:.2f}%")
            row_kv("% bases >=100x", f"{float(hs.get('pct_target_bases_100x',0))*100:.2f}%")
            row_kv("% bases >=250x", f"{float(hs.get('pct_target_bases_250x',0))*100:.2f}%")
            row_kv("% bases >=500x", f"{float(hs.get('pct_target_bases_500x',0))*100:.2f}%")
            row_kv("% bases >=1000x", f"{float(hs.get('pct_target_bases_1000x',0))*100:.2f}%")
            row_kv("On-target bases", hs.get("on_target_bases",""))
            row_kv("Total reads", hs.get("total_reads",""))
            row_kv("PF unique reads", hs.get("pf_unique_reads",""))

    # Mismatch rate
    mm = d.get("mismatch_rate")
    if mm:
        h("Mismatch Rate")
        for k, v in mm.items():
            row_kv(k, v)

    lines.append("")
    lines.append("=" * 60)
    lines.append("  End of Report")
    lines.append("=" * 60)
    lines.append("")

    txt = "\n".join(lines)
    sample = order["sample_name"]
    filename = f"{sample}_QC_report.txt"
    from flask import Response
    return Response(
        txt,
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.route("/api/orders/<order_id>/vcf_data")
def api_vcf_data(order_id):
    """Parse the VariantsToTable annotated txt and return JSON for the review table."""
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"error": "Order not found"}), 404
    order = dict_from_row(row)
    sample = order["sample_name"]

    candidates = glob.glob(
        os.path.join(RESULTS_DIR, sample, "**", f"{sample}*annotated*vcf.txt"),
        recursive=True,
    )
    if not candidates:
        return jsonify({"error": "No annotated VCF table found"}), 404
    fp = candidates[0]

    display_cols = [
        "CHROM", "POS", "REF", "ALT", "GENE", "TRANSCRIPT", "EFFECT", "IMPACT",
        "HGVS_C", "HGVS_P", "HGVS_G", "AF", "DP", "VD", "HIAF", "MQ", "GT",
        "DUPRATE", "FILTER", "TYPE", "QUAL", "ID", "DBSNP_COMMON",
    ]

    def parse_ann(ann_str):
        """Extract gene, effect, impact, transcript, HGVS.c, HGVS.p from SnpEff ANN.

        Prefers the first entry whose transcript ID starts with 'NM_'.
        Falls back to the very first entry if no NM_ transcript is found.
        """
        if not ann_str or ann_str in ("NA", "."):
            return "", "", "", "", "", ""
        entries = ann_str.split(",")
        chosen = entries[0]
        for entry in entries:
            p = entry.split("|")
            tx = p[6] if len(p) > 6 else ""
            if tx.startswith("NM_"):
                chosen = entry
                break
        parts = chosen.split("|")
        effect     = parts[1]  if len(parts) > 1  else ""
        impact     = parts[2]  if len(parts) > 2  else ""
        gene       = parts[3]  if len(parts) > 3  else ""
        transcript = parts[6]  if len(parts) > 6  else ""
        hgvs_c     = parts[9]  if len(parts) > 9  else ""
        hgvs_p     = parts[10] if len(parts) > 10 else ""
        return gene, effect, impact, transcript, hgvs_c, hgvs_p

    try:
        order_ref = order.get("reference") or "hg38"
        by_position, by_cnumber, by_gene = _build_list_lookup(reference=order_ref)
        rows_out = []
        blacklisted = 0
        with open(fp, "r") as fh:
            header = fh.readline().rstrip("\n").split("\t")
            col_idx = {c: i for i, c in enumerate(header)}
            # Genotype fields are emitted as "<sample>.GT" etc — find them
            gt_col = next((c for c in col_idx if c.endswith(".GT")), None)
            for line in fh:
                vals = line.rstrip("\n").split("\t")
                gene, effect, impact, transcript, hgvs_c, hgvs_p = parse_ann(
                    vals[col_idx["ANN"]] if "ANN" in col_idx else ""
                )
                rec = {}
                for c in display_cols:
                    if c == "GENE":
                        rec[c] = gene
                    elif c == "TRANSCRIPT":
                        rec[c] = transcript
                    elif c == "EFFECT":
                        rec[c] = effect
                    elif c == "IMPACT":
                        rec[c] = impact
                    elif c == "HGVS_C":
                        rec[c] = hgvs_c
                    elif c == "HGVS_P":
                        rec[c] = hgvs_p
                    elif c == "HGVS_G":
                        chrom_val = vals[col_idx["CHROM"]] if "CHROM" in col_idx else ""
                        pos_val   = vals[col_idx["POS"]]   if "POS"   in col_idx else ""
                        ref_val   = vals[col_idx["REF"]]   if "REF"   in col_idx else ""
                        alt_val   = vals[col_idx["ALT"]]   if "ALT"   in col_idx else ""
                        nc = _NC_HG38.get(chrom_val, "")
                        rec[c] = f"{nc}:g.{pos_val}{ref_val}>{alt_val}" if nc and pos_val else ""
                    elif c == "GT":
                        rec[c] = vals[col_idx[gt_col]] if gt_col else ""
                    elif c in col_idx:
                        rec[c] = vals[col_idx[c]]
                    else:
                        rec[c] = ""
                # Annotate with variant-list membership
                tag = _get_list_tag(
                    rec.get("CHROM", ""),
                    rec.get("POS", ""),
                    rec.get("REF", ""),
                    rec.get("ALT", ""),
                    rec.get("ID", ""),
                    by_position, by_cnumber, by_gene,
                    gene=rec.get("GENE", ""),
                    hgvs_c=rec.get("HGVS.C", ""),
                )
                if tag == "blacklist":
                    blacklisted += 1
                rec["_list_tag"] = tag  # '' | 'whitelist' | 'onhold' | 'blacklist'
                rows_out.append(rec)
        return jsonify({
            "success": True,
            "columns": display_cols,
            "data": rows_out,
            "sample_name": sample,
            "order_name": order.get("order_name", ""),
            "total": len(rows_out),
            "blacklisted_count": blacklisted,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# Routes — IGV (BAM viewer)
# ---------------------------------------------------------------------------
def _igv_disabled_response():
    return jsonify({"error": "IGV feature disabled"}), 404


def _order_results_root(order):
    """Return the absolute, real path of results/<sample>/ for this order (or None)."""
    sample = (order or {}).get("sample_name", "")
    if not sample:
        return None
    root = os.path.join(RESULTS_DIR, sample)
    if not os.path.isdir(root):
        return None
    return os.path.realpath(root)


def _safe_resolve_under(root_real, rel_path):
    """Resolve rel_path against root_real, ensuring no traversal. Returns abs path or None."""
    if not root_real or not rel_path:
        return None
    if rel_path.startswith("/") or ".." in rel_path.replace("\\", "/").split("/"):
        return None
    candidate = os.path.realpath(os.path.join(root_real, rel_path))
    if not (candidate == root_real or candidate.startswith(root_real + os.sep)):
        return None
    return candidate


# Heuristic: tokens that mark a BAM as "ancillary" (paraphase, SMN etc.).
# Kept empty for Roche_nxt — every produced BAM is the main exome/panel track.
_ANCILLARY_TOKENS = ()


def _bam_label_sort_key(name):
    """Prefer output/bam final BAM > QC_report rmdups BAM > others; newest mtime first."""
    lower = name.lower()
    if "/output/bam/" in lower or lower.endswith("_sorted.bam"):
        rank = 0
    elif "_sorted_rmdups.bam" in lower:
        rank = 1
    else:
        rank = 2
    return rank


def _find_bai_for(bam_abs):
    """Return absolute path of the BAI file for bam_abs, or None."""
    for bai in (bam_abs + ".bai", bam_abs[:-4] + ".bai"):
        if os.path.isfile(bai):
            return bai
    return None


def _try_index_bam(bam_abs):
    """Best-effort: run `samtools index` inside the analysis image to produce a BAI."""
    try:
        host_root = HOST_DIR
        bam_real = os.path.realpath(bam_abs)
        results_real = os.path.realpath(RESULTS_DIR)
        if not bam_real.startswith(results_real + os.sep):
            return None
        # Translate container path (/roche_nxt/...) to host path for the sibling docker run.
        rel_from_base = os.path.relpath(bam_real, os.path.realpath(BASE_DIR))
        host_bam = os.path.join(host_root, rel_from_base)
        uid = subprocess.run(["id", "-u"], capture_output=True, text=True).stdout.strip()
        gid = subprocess.run(["id", "-g"], capture_output=True, text=True).stdout.strip()
        subprocess.run(
            [
                "docker", "run", "--rm",
                "--user", f"{uid}:{gid}",
                "-v", f"{host_root}:/work_nxt",
                ANALYSIS_IMAGE,
                "samtools", "index",
                f"/work_nxt/{rel_from_base}",
            ],
            capture_output=True, text=True, timeout=180,
        )
    except Exception:
        return None
    return _find_bai_for(bam_abs)


def _list_bam_tracks(order, auto_index=True):
    """Return (tracks, results_root_real). tracks: list of dicts."""
    root_real = _order_results_root(order)
    if not root_real:
        return [], None
    sample = order["sample_name"]
    bams = []
    for bam_abs in glob.glob(os.path.join(root_real, "**", "*.bam"), recursive=True):
        bam_real = os.path.realpath(bam_abs)
        if not bam_real.startswith(root_real + os.sep):
            continue
        bams.append(bam_real)
    bams = sorted(set(bams), key=lambda p: (_bam_label_sort_key(p), -os.path.getmtime(p)))

    tracks = []
    for bam_abs in bams:
        bai_abs = _find_bai_for(bam_abs)
        if not bai_abs and auto_index:
            bai_abs = _try_index_bam(bam_abs)
        rel_bam = os.path.relpath(bam_abs, root_real).replace(os.sep, "/")
        rel_bai = os.path.relpath(bai_abs, root_real).replace(os.sep, "/") if bai_abs else None
        name = os.path.basename(bam_abs)
        ancillary = any(tok in name.lower() for tok in _ANCILLARY_TOKENS)
        tracks.append({
            "rel_path": rel_bam,
            "label": name,
            "has_index": bool(bai_abs),
            "index_rel_path": rel_bai,
            "size": os.path.getsize(bam_abs),
            "mtime": os.path.getmtime(bam_abs),
            "ancillary": ancillary,
        })
    return tracks, root_real


@app.route("/api/orders/<order_id>/coverage-context")
def api_coverage_context(order_id):
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"error": "Order not found"}), 404
    order = dict_from_row(row)
    tracks, root_real = _list_bam_tracks(order, auto_index=True)
    indexed = [t for t in tracks if t["has_index"] and not t["ancillary"]]
    primary = indexed[0] if indexed else None
    return jsonify({
        "success": True,
        "order_id": order_id,
        "sample_name": order.get("sample_name", ""),
        "order_name": order.get("order_name", ""),
        "genome_id": order.get("reference", "hg38") or "hg38",
        "bam_tracks": tracks,
        "primary_track": primary,
        "interpretation_genes": [],
        "results_root_exists": bool(root_real),
    })


@app.route("/api/orders/<order_id>/index_bam", methods=["POST"])
def api_index_bam(order_id):
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "error": "Order not found"}), 404
    order = dict_from_row(row)
    rel = (request.json or {}).get("rel_path", "") if request.is_json else request.form.get("rel_path", "")
    root_real = _order_results_root(order)
    abs_bam = _safe_resolve_under(root_real, rel)
    if not abs_bam or not os.path.isfile(abs_bam) or not abs_bam.endswith(".bam"):
        return jsonify({"success": False, "error": "BAM not found"}), 404
    bai_abs = _try_index_bam(abs_bam)
    if not bai_abs:
        return jsonify({"success": False, "error": "samtools index failed"}), 500
    return jsonify({
        "success": True,
        "index_rel_path": os.path.relpath(bai_abs, root_real).replace(os.sep, "/"),
    })


def _serve_range(abs_path, mimetype):
    """Serve abs_path with HTTP Range support (IGV.js requires this for BAM)."""
    try:
        total = os.path.getsize(abs_path)
    except OSError:
        abort(404)

    range_header = request.headers.get("Range", "").strip()
    if not range_header:
        resp = send_file(abs_path, mimetype=mimetype, conditional=True)
        resp.headers["Accept-Ranges"] = "bytes"
        resp.headers["Content-Length"] = str(total)
        return resp

    # "bytes=<start>-<end>"
    if not range_header.lower().startswith("bytes="):
        abort(416)
    try:
        spec = range_header.split("=", 1)[1].split(",", 1)[0].strip()
        start_str, end_str = spec.split("-", 1)
        start = int(start_str) if start_str else 0
        end = int(end_str) if end_str else total - 1
    except ValueError:
        abort(416)
    if start < 0 or end >= total or start > end:
        abort(416)

    length = end - start + 1
    chunk_size = 1024 * 1024

    def _gen():
        with open(abs_path, "rb") as f:
            f.seek(start)
            remaining = length
            while remaining > 0:
                data = f.read(min(chunk_size, remaining))
                if not data:
                    break
                remaining -= len(data)
                yield data

    resp = Response(_gen(), status=206, mimetype=mimetype, direct_passthrough=True)
    resp.headers["Content-Range"] = f"bytes {start}-{end}/{total}"
    resp.headers["Accept-Ranges"] = "bytes"
    resp.headers["Content-Length"] = str(length)
    return resp


def _guess_mimetype(path):
    lower = path.lower()
    if lower.endswith(".bam"):
        return "application/octet-stream"
    if lower.endswith(".bai") or lower.endswith(".bam.bai"):
        return "application/octet-stream"
    return "application/octet-stream"


@app.route("/api/orders/<order_id>/file/<path:filename>", methods=["GET", "HEAD"])
def api_order_file(order_id, filename):
    """Serve result files for IGV.js. Only BAM/BAI are allowed by default."""
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"error": "Order not found"}), 404
    order = dict_from_row(row)
    root_real = _order_results_root(order)
    abs_path = _safe_resolve_under(root_real, filename)
    if not abs_path or not os.path.isfile(abs_path):
        return jsonify({"error": "File not found"}), 404
    # Allowlist: only alignment tracks via this endpoint.
    lower = abs_path.lower()
    if not (lower.endswith(".bam") or lower.endswith(".bai") or lower.endswith(".bam.bai")):
        return jsonify({"error": "File type not allowed"}), 403

    mimetype = _guess_mimetype(abs_path)
    try:
        total = os.path.getsize(abs_path)
    except OSError:
        return jsonify({"error": "File not found"}), 404

    if request.method == "HEAD":
        resp = make_response("", 200)
        resp.headers["Content-Type"] = mimetype
        resp.headers["Content-Length"] = str(total)
        resp.headers["Accept-Ranges"] = "bytes"
        return resp

    return _serve_range(abs_path, mimetype)


# ---------------------------------------------------------------------------
# Routes — Liftover (hg38 -> hg19, display-only)
# ---------------------------------------------------------------------------
# Lazy-loaded pyliftover instance. The chain file is large-ish (~1MB gzipped)
# and parsing takes ~0.5s, so we cache the parsed LiftOver object.
_LIFTOVER_LOCK_KEY = "_liftover_hg38_to_hg19"


def _get_liftover_hg38_to_hg19():
    """Return a cached pyliftover.LiftOver instance, or None if unavailable."""
    cached = app.config.get(_LIFTOVER_LOCK_KEY)
    if cached is not None:
        # (loader_result, path_mtime) tuple — invalidate if file changed.
        lo, loaded_mtime = cached
        try:
            current_mtime = os.path.getmtime(LIFTOVER_CHAIN_HG38_TO_HG19)
        except OSError:
            current_mtime = None
        if current_mtime == loaded_mtime:
            return lo

    if not ENABLE_HG19_VIEW:
        return None
    if not os.path.isfile(LIFTOVER_CHAIN_HG38_TO_HG19):
        app.config[_LIFTOVER_LOCK_KEY] = (None, None)
        return None
    try:
        from pyliftover import LiftOver  # type: ignore
    except ImportError:
        app.logger.error("pyliftover not installed; hg19 view disabled.")
        app.config[_LIFTOVER_LOCK_KEY] = (None, None)
        return None
    try:
        lo = LiftOver(LIFTOVER_CHAIN_HG38_TO_HG19)
        app.config[_LIFTOVER_LOCK_KEY] = (lo, os.path.getmtime(LIFTOVER_CHAIN_HG38_TO_HG19))
        return lo
    except Exception as exc:
        app.logger.error("Failed to load liftover chain %s: %s", LIFTOVER_CHAIN_HG38_TO_HG19, exc)
        app.config[_LIFTOVER_LOCK_KEY] = (None, None)
        return None


def _lift_one_hg38_to_hg19(lo, chrom, pos_1based):
    """Lift a single 1-based hg38 coordinate to hg19. Returns (chrom19, pos19) or None."""
    if lo is None or not chrom:
        return None
    try:
        pos0 = int(pos_1based) - 1  # pyliftover is 0-based
    except (TypeError, ValueError):
        return None
    if pos0 < 0:
        return None
    try:
        hits = lo.convert_coordinate(chrom, pos0)
    except Exception:
        return None
    if not hits:
        return None
    # Prefer matches on the forward strand of the mapped chrom; otherwise first hit.
    forward = [h for h in hits if len(h) >= 3 and h[2] == "+"]
    h = (forward or hits)[0]
    chrom19, pos0_19 = h[0], h[1]
    return (chrom19, pos0_19 + 1)


@app.route("/api/liftover", methods=["POST"])
def api_liftover():
    """Batch lift hg38 coordinates to hg19 for display purposes.

    Request JSON:  { "positions": [{"chrom": "chr1", "pos": 12345}, ...] }
    Response JSON: {
        "success": true,
        "assembly_from": "hg38",
        "assembly_to": "hg19",
        "results": [
            {"chrom": "chr1", "pos": 12345, "chrom19": "chr1", "pos19": 12300, "ok": true},
            ...
        ]
    }
    Coordinates for which liftover fails are returned with ok=false.
    """
    if not ENABLE_HG19_VIEW:
        return jsonify({"error": "hg19 view disabled"}), 404

    lo = _get_liftover_hg38_to_hg19()
    if lo is None:
        return jsonify({
            "error": "Liftover chain not available. "
                     "Check ENABLE_HG19_VIEW and LIFTOVER_CHAIN_HG38_TO_HG19."
        }), 503

    payload = request.get_json(silent=True) or {}
    positions = payload.get("positions") or []
    if not isinstance(positions, list):
        return jsonify({"error": "'positions' must be a list"}), 400
    # Guardrail: avoid accidental huge requests (a typical exome review has <10k rows).
    if len(positions) > 200000:
        return jsonify({"error": "Too many positions (>200k)"}), 400

    out = []
    for entry in positions:
        if not isinstance(entry, dict):
            out.append({"ok": False})
            continue
        chrom = str(entry.get("chrom") or "").strip()
        pos_raw = entry.get("pos")
        try:
            pos = int(pos_raw)
        except (TypeError, ValueError):
            out.append({"chrom": chrom, "pos": pos_raw, "ok": False})
            continue
        mapped = _lift_one_hg38_to_hg19(lo, chrom, pos)
        if mapped is None:
            out.append({"chrom": chrom, "pos": pos, "ok": False})
        else:
            out.append({
                "chrom": chrom,
                "pos": pos,
                "chrom19": mapped[0],
                "pos19": mapped[1],
                "ok": True,
            })

    return jsonify({
        "success": True,
        "assembly_from": "hg38",
        "assembly_to": "hg19",
        "results": out,
    })


# ---------------------------------------------------------------------------
# Routes — Longitudinal VAF comparison
# ---------------------------------------------------------------------------
@app.route("/api/orders/<order_id>/longitudinal_data")
def api_longitudinal_data(order_id):
    """Build a VAF comparison table across the current followup + any prior followups."""
    db = get_db()
    row = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not row:
        return jsonify({"error": "Order not found"}), 404
    order = dict_from_row(row)
    if order.get("order_type") != "longitudinal":
        return jsonify({"error": "Not a longitudinal order"}), 400

    def _parse_longit_csv(filepath):
        if not os.path.isfile(filepath):
            return None
        rows = []
        with open(filepath, "r", errors="replace") as fh:
            reader = csv.DictReader(fh)
            for r in reader:
                rows.append(r)
        return rows

    sample = order["sample_name"]
    longit_candidates = glob.glob(
        os.path.join(RESULTS_DIR, sample, "**", f"{sample}_longitudinal_analysis.csv"),
        recursive=True,
    )
    current_data = _parse_longit_csv(longit_candidates[0]) if longit_candidates else None

    followup_ids = [fid.strip() for fid in (order.get("followup_order_ids") or "").split(",") if fid.strip()]
    prior_followups = []
    for fid in followup_ids:
        frow = db.execute("SELECT sample_name, order_name, created_at FROM orders WHERE id=?", (fid,)).fetchone()
        if not frow:
            continue
        fsample = frow["sample_name"]
        fc = glob.glob(
            os.path.join(RESULTS_DIR, fsample, "**", f"{fsample}_longitudinal_analysis.csv"),
            recursive=True,
        )
        fdata = _parse_longit_csv(fc[0]) if fc else None
        if fdata:
            prior_followups.append({
                "order_id": fid,
                "order_name": frow["order_name"],
                "sample_name": fsample,
                "created_at": frow["created_at"],
                "data": fdata,
            })

    baseline_id = order.get("baseline_order_id", "")
    baseline_info = None
    if baseline_id:
        brow = db.execute("SELECT order_name, sample_name FROM orders WHERE id=?", (baseline_id,)).fetchone()
        if brow:
            baseline_info = {"order_id": baseline_id, "order_name": brow["order_name"], "sample_name": brow["sample_name"]}

    return jsonify({
        "success": True,
        "order_id": order_id,
        "order_name": order.get("order_name", ""),
        "sample_name": sample,
        "baseline": baseline_info,
        "current_followup": {
            "order_id": order_id,
            "order_name": order.get("order_name", ""),
            "sample_name": sample,
            "created_at": order.get("created_at", ""),
            "data": current_data,
        },
        "prior_followups": prior_followups,
    })


# ---------------------------------------------------------------------------
# .env file helpers
# ---------------------------------------------------------------------------
ENV_FILE = os.path.join(BASE_DIR, ".env")


def read_env_file():
    """Read .env file into a dict."""
    env = {}
    if os.path.isfile(ENV_FILE):
        with open(ENV_FILE, "r") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    env[k.strip()] = v.strip()
    return env


def write_env_file(env):
    """Write dict back to .env file, preserving key order."""
    with open(ENV_FILE, "w") as f:
        for k, v in env.items():
            f.write(f"{k}={v}\n")


# ---------------------------------------------------------------------------
# Routes — Settings API
# ---------------------------------------------------------------------------
@app.route("/api/settings")
def api_get_settings():
    settings = get_all_settings()
    env = read_env_file()
    settings["fastq_host_dir"] = env.get("FASTQ_HOST_DIR", FASTQ_HOST_DIR)
    settings["bed_host_dir"] = env.get("BED_HOST_DIR", BED_HOST_DIR)
    settings["max_cpus"] = env.get("MAX_CPUS", "0")
    settings["max_memory"] = env.get("MAX_MEMORY", "0")
    settings["max_concurrent_samples"] = env.get("MAX_CONCURRENT_SAMPLES", "0")
    _def_ref = settings.get("default_reference", "hg38")
    settings["default_reference"] = _def_ref if _def_ref in ("hg38", "hg19") else "hg38"

    # Longitudinal defaults (DB values take precedence)
    _sr_defaults = {
        "sr_germline_cutoff": "0.005",
        "sr_min_af":          "0.005",
        "sr_max_af":          "0.35",
        "sr_min_dp":          "1000",
        "sr_min_vd":          "15",
        "sr_min_mq":          "55",
        "sr_min_qual":        "45",
        "sr_min_sbf":         "1e-05",
        "sr_max_nm":          "4",
    }
    _la_defaults = {
        "la_reads_threshold":  "1000",
        "la_pvalue_threshold": "0.001",
        "la_vaf_threshold":    "0.1",
        "la_n_sim":            "10000",
        "la_blist_type":       "variant",
    }
    # Baseline defaults
    _bl_defaults = {
        "enable_umi":             "true",
        "umi_read_structure":     "3M3S+T 3M3S+T",
        "seqtk_sample_size":      "40000000",
        "seqtk_seed":             "12345",
        "enable_subsampling":     "false",
        "subsample_threshold_gb": "20",
        "fastp_options":          "-g -W 5 -q 20 -u 40 -x -3 -l 75 -c",
        "min_reads":              "1",
        "min_base_quality":       "20",
        "max_read_error_rate":    "0.025",
        "max_base_error_rate":    "0.1",
        "max_no_call_fraction":   "0.1",
    }
    for k, v in {**_sr_defaults, **_la_defaults, **_bl_defaults}.items():
        settings.setdefault(k, v)
    settings.setdefault("longitudinal_enabled", "true")
    settings.setdefault("panels_enabled", "exome")
    # RNA pipeline reference paths (empty = auto-resolve from nextflow.config genomes block)
    settings.setdefault("rna_star_index", "")
    settings.setdefault("rna_gtf", "")
    settings.setdefault("rna_bed12", "")
    settings.setdefault("rna_fastp_options", "-g -W 5 -q 20 -u 40 -x -3 -l 50 -c")
    settings.setdefault("rna_ctat_lib", "")

    return jsonify(settings)


@app.route("/api/settings", methods=["POST"])
def api_save_settings():
    if not is_admin_user(session.get("user_id")):
        return jsonify({"success": False, "error": "관리자만 설정을 변경할 수 있습니다."}), 403
    data = request.json
    db = get_db()

    restart_needed = False
    env = None

    if "fastq_host_dir" in data:
        new_dir = data.pop("fastq_host_dir").strip()
        if new_dir:
            env = env or read_env_file()
            env["FASTQ_HOST_DIR"] = new_dir
            # Update running process immediately so new analysis runs use the new path
            os.environ["FASTQ_HOST_DIR"] = new_dir
            global FASTQ_HOST_DIR
            FASTQ_HOST_DIR = new_dir
            restart_needed = True

    if "bed_host_dir" in data:
        new_dir = data.pop("bed_host_dir").strip()
        if new_dir:
            env = env or read_env_file()
            env["BED_HOST_DIR"] = new_dir
            os.environ["BED_HOST_DIR"] = new_dir
            global BED_HOST_DIR
            BED_HOST_DIR = new_dir
            restart_needed = True

    for env_key in ["max_cpus", "max_memory", "max_concurrent_samples"]:
        if env_key in data:
            val = str(data.pop(env_key)).strip()
            env = env or read_env_file()
            env[env_key.upper()] = val
            os.environ[env_key.upper()] = val
            restart_needed = True

    if env is not None:
        write_env_file(env)

    if "default_reference" in data:
        data["default_reference"] = data["default_reference"] if data["default_reference"] in ("hg38", "hg19") else "hg38"

    for k, v in data.items():
        db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (k, str(v)))
    db.commit()
    return jsonify({"success": True, "restart_needed": restart_needed})


@app.route("/api/restart", methods=["POST"])
def api_restart():
    """Restart the web container via Docker socket (async — caller must poll /api/health)."""
    import threading

    def _do_restart():
        import time
        time.sleep(0.5)  # Allow HTTP response to be sent first
        try:
            subprocess.run(
                ["docker", "restart", "roche_nxt_web"],
                timeout=30,
                capture_output=True,
            )
        except Exception:
            pass

    t = threading.Thread(target=_do_restart, daemon=True)
    t.start()
    return jsonify({"success": True, "message": "재시작 요청이 전송되었습니다."})


@app.route("/api/health", methods=["GET"])
def api_health():
    return jsonify({"status": "ok"})


# ---------------------------------------------------------------------------
# Variant Lists (Blacklist / Whitelist / On-hold)
# ---------------------------------------------------------------------------

_VL_PRIORITY = {"blacklist": 3, "whitelist": 2, "onhold": 1}

# Chromosome → RefSeq NC accession for HGVS.g construction
_NC_HG38 = {
    "chr1":"NC_000001.11","chr2":"NC_000002.12","chr3":"NC_000003.12",
    "chr4":"NC_000004.12","chr5":"NC_000005.10","chr6":"NC_000006.12",
    "chr7":"NC_000007.14","chr8":"NC_000008.11","chr9":"NC_000009.12",
    "chr10":"NC_000010.11","chr11":"NC_000011.10","chr12":"NC_000012.12",
    "chr13":"NC_000013.11","chr14":"NC_000014.9","chr15":"NC_000015.10",
    "chr16":"NC_000016.10","chr17":"NC_000017.11","chr18":"NC_000018.10",
    "chr19":"NC_000019.10","chr20":"NC_000020.11","chr21":"NC_000021.9",
    "chr22":"NC_000022.11","chrX":"NC_000023.11","chrY":"NC_000024.10",
    "chrM":"NC_012920.1",
}


def _norm_chrom(c):
    """Normalize chromosome name for comparison (add 'chr' prefix if absent)."""
    c = (c or "").strip()
    if c and not c.lower().startswith("chr"):
        c = "chr" + c
    return c.lower()


def _build_list_lookup(reference="hg38"):
    """Return (by_position, by_cnumber, by_gene) lookup dicts for fast matching.

    Position entries are filtered to genome == reference OR genome == 'any'.
    cnumber/gene entries always match regardless of reference.
    """
    db = get_db()
    rows = db.execute(
        "SELECT list_type, entry_type, cnumber, chrom, pos, ref, alt, gene, genome FROM variant_lists"
    ).fetchall()

    by_position = {}  # (chrom_norm, pos, ref_u, alt_u) or (chrom_norm, pos, '', '') -> list_type
    by_cnumber  = {}  # cnumber_lower -> list_type
    by_gene     = {}  # GENE_UPPER -> list_type  (gene-level entries)

    for r in rows:
        r = dict_from_row(r)
        lt = r["list_type"]
        p  = _VL_PRIORITY.get(lt, 0)
        et = r.get("entry_type", "")
        genome = (r.get("genome") or "hg38").strip().lower()

        if et == "gene":
            gene_u = (r.get("gene") or "").strip().upper()
            if gene_u:
                if gene_u not in by_gene or p > _VL_PRIORITY.get(by_gene[gene_u], 0):
                    by_gene[gene_u] = lt
        elif et == "cnumber" and (r.get("cnumber") or "").strip():
            key = r["cnumber"].strip().lower()
            if key not in by_cnumber or p > _VL_PRIORITY.get(by_cnumber[key], 0):
                by_cnumber[key] = lt
        else:
            # Position entries: only match if genome matches or 'any'
            if genome not in (reference.lower(), "any"):
                continue
            chrom = _norm_chrom(r.get("chrom") or "")
            pos   = int(r.get("pos") or 0)
            if not chrom or not pos:
                continue
            ref_u = (r.get("ref") or "").strip().upper()
            alt_u = (r.get("alt") or "").strip().upper()
            for key in [(chrom, pos, ref_u, alt_u), (chrom, pos, "", "")]:
                if key not in by_position or p > _VL_PRIORITY.get(by_position[key], 0):
                    by_position[key] = lt

    return by_position, by_cnumber, by_gene


def _get_list_tag(chrom, pos_str, ref, alt, vcf_id, by_position, by_cnumber, by_gene=None, gene="", hgvs_c=""):
    """Return 'blacklist', 'whitelist', 'onhold', or '' for this variant.

    Matching (highest priority wins):
      1. Genomic position  (CHR + POS [+ REF + ALT])
      2. HGVS.c            (c.Number from annotation, or VCF ID)
      3. Gene-level        (all variants of a gene)
    """
    best = ""
    best_p = 0
    by_gene = by_gene or {}

    chrom_n = _norm_chrom(chrom)
    pos = int(pos_str or 0)
    ref_u = (ref or "").strip().upper()
    alt_u = (alt or "").strip().upper()

    # 1. Position match
    for key in [(chrom_n, pos, ref_u, alt_u), (chrom_n, pos, "", "")]:
        lt = by_position.get(key)
        if lt:
            p = _VL_PRIORITY.get(lt, 0)
            if p > best_p:
                best, best_p = lt, p

    # 2. HGVS.c / VCF-ID match
    for cand in [hgvs_c, vcf_id]:
        if not cand or cand in (".", ""):
            continue
        for tok in str(cand).split(";"):
            tok = tok.strip().lower()
            if tok and tok in by_cnumber:
                p = _VL_PRIORITY.get(by_cnumber[tok], 0)
                if p > best_p:
                    best, best_p = by_cnumber[tok], p

    # 3. Gene-level match
    if by_gene and gene:
        gene_u = gene.strip().upper()
        lt = by_gene.get(gene_u)
        if lt:
            p = _VL_PRIORITY.get(lt, 0)
            if p > best_p:
                best, best_p = lt, p

    return best


def _parse_variant_list_file(file_bytes, filename, list_type, skip_header=False):
    """Parse an uploaded xlsx / txt / bed file and return (entries, skipped_rows).

    skip_header: when True, forcibly treat the first row as a header row
                 (ignores auto-detection and always skips row 0).
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "txt"
    now = datetime.now().isoformat()
    entries = []
    skipped_rows = []

    if ext in ("xlsx", "xls"):
        import io
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl가 설치되어 있지 않습니다 (pip install openpyxl).")
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        raw_rows = [
            [str(cell or "").strip() for cell in row]
            for row in ws.iter_rows(values_only=True)
        ]
        wb.close()
    elif ext == "bed":
        lines = [
            l.rstrip() for l in file_bytes.decode("utf-8", errors="replace").splitlines()
            if l.strip() and not l.startswith("#")
            and not l.lower().startswith("track") and not l.lower().startswith("browser")
        ]
        for line in lines:
            cols = line.split("\t") if "\t" in line else line.split()
            if len(cols) < 2:
                continue
            chrom = cols[0].strip()
            try:
                pos = int(cols[1]) + 1  # BED 0-based start → 1-based
            except Exception:
                continue
            note = cols[3].strip() if len(cols) > 3 else ""
            entries.append({
                "id": str(uuid.uuid4()), "list_type": list_type,
                "entry_type": "position", "cnumber": "",
                "chrom": chrom, "pos": pos, "ref": "", "alt": "",
                "gene": "", "note": note, "created_at": now, "updated_at": now,
            })
        return entries, []
    else:
        content = file_bytes.decode("utf-8", errors="replace")
        lines = [l.rstrip() for l in content.splitlines() if l.strip() and not l.startswith("#")]
        if not lines:
            return [], []
        sep = "\t" if "\t" in lines[0] else ","
        raw_rows = [l.split(sep) for l in lines]

    if not raw_rows:
        return []

    # Detect or force header row
    HEADER_KEYWORDS = {
        "CNUMBER", "C_NUMBER", "C#", "CNUMBER.", "VARIANT_ID", "VARIANT-ID",
        "VARIANT", "CHROM", "CHR", "CHROMOSOME", "POS", "POSITION", "START",
        "REF", "ALT", "GENE", "GENE_NAME", "NOTE", "COMMENT", "ID",
        "분류", "CATEGORY", "비율", "RATIO", "빈도", "FREQUENCY",
    }

    # Regex to detect position-format variants: e.g. "11-69532129-G-T-snv"
    _POS_VARIANT_RE = re.compile(
        r'^([A-Za-z0-9]+)-(\d+)-([A-Za-z*]+)-([A-Za-z*]+)(?:-\w+)?$'
    )

    def _parse_variant_field(v):
        """Return (entry_type, chrom, pos_int, ref, alt, cnumber) from a Variant field."""
        v = (v or "").strip()
        if not v:
            return None
        m = _POS_VARIANT_RE.match(v)
        if m:
            chrom_raw, pos_str, ref, alt = m.group(1), m.group(2), m.group(3).upper(), m.group(4).upper()
            chrom = chrom_raw if chrom_raw.lower().startswith("chr") else f"chr{chrom_raw}"
            return ("position", chrom, int(pos_str), ref, alt, "")
        return ("cnumber", "", 0, "", "", v)
    first_upper = [c.strip().upper().replace(" ", "_") for c in raw_rows[0]]

    if skip_header:
        has_header = True
    else:
        has_header = any(c in HEADER_KEYWORDS for c in first_upper)

    if has_header:
        header = first_upper
        data_rows = raw_rows[1:]
    else:
        header = None
        data_rows = raw_rows

    def _gi(h, *names):
        for n in names:
            if n in h:
                return h.index(n)
        return -1

    def _gv(row, i):
        return row[i].strip() if i >= 0 and i < len(row) else ""

    if header is not None:
        cn_col  = _gi(header, "CNUMBER", "C_NUMBER", "C#", "CNUMBER.", "VARIANT_ID", "VARIANT-ID", "VARIANT", "ID")
        chr_col = _gi(header, "CHROM", "CHR", "CHROMOSOME")
        pos_col = _gi(header, "POS", "POSITION", "START")
        ref_col = _gi(header, "REF", "REFERENCE")
        alt_col = _gi(header, "ALT", "ALTERNATE")
        gn_col  = _gi(header, "GENE", "GENE_NAME")
        nt_col  = _gi(header, "NOTE", "COMMENT", "DESCRIPTION", "REMARK", "비고")
        cat_col = _gi(header, "분류", "CATEGORY", "CLASS", "TYPE")
        rat_col = _gi(header, "비율", "RATIO")
        freq_col = _gi(header, "빈도", "FREQUENCY", "FREQ")

        skipped_rows = []
        for row_idx, row in enumerate(data_rows, start=2):  # 2 = first data row (row 1 = header)
            if not any(v.strip() for v in row):
                continue  # truly empty row
            cn  = _gv(row, cn_col)
            chm = _gv(row, chr_col)
            ps  = _gv(row, pos_col)
            cat = _gv(row, cat_col)
            rat = _gv(row, rat_col)
            freq = _gv(row, freq_col)
            gene = _gv(row, gn_col)
            note = _gv(row, nt_col)

            # Try to parse Variant column as position format first
            parsed = _parse_variant_field(cn) if cn else None

            if parsed and parsed[0] == "position":
                _, chm, pos_int, ref, alt, _ = parsed
                entries.append({
                    "id": str(uuid.uuid4()), "list_type": list_type,
                    "entry_type": "position", "cnumber": "",
                    "chrom": chm, "pos": pos_int,
                    "ref": ref, "alt": alt,
                    "gene": gene, "note": note,
                    "category": cat, "ratio": rat or freq,
                    "created_at": now, "updated_at": now,
                })
            elif chm and ps:
                try:
                    pos_int = int(float(ps))
                except Exception:
                    pos_int = 0
                entries.append({
                    "id": str(uuid.uuid4()), "list_type": list_type,
                    "entry_type": "position", "cnumber": cn,
                    "chrom": chm, "pos": pos_int,
                    "ref": _gv(row, ref_col).upper(), "alt": _gv(row, alt_col).upper(),
                    "gene": gene, "note": note,
                    "category": cat, "ratio": rat or freq,
                    "created_at": now, "updated_at": now,
                })
            elif cn:
                entries.append({
                    "id": str(uuid.uuid4()), "list_type": list_type,
                    "entry_type": "cnumber", "cnumber": cn,
                    "chrom": chm, "pos": 0,
                    "ref": _gv(row, ref_col).upper(), "alt": _gv(row, alt_col).upper(),
                    "gene": gene, "note": note,
                    "category": cat, "ratio": rat or freq,
                    "created_at": now, "updated_at": now,
                })
            elif gene:
                # Gene-only row: match all variants from this gene
                entries.append({
                    "id": str(uuid.uuid4()), "list_type": list_type,
                    "entry_type": "gene", "cnumber": "",
                    "chrom": "", "pos": 0, "ref": "", "alt": "",
                    "gene": gene, "note": note,
                    "category": cat, "ratio": rat or freq,
                    "created_at": now, "updated_at": now,
                })
            else:
                skipped_rows.append({"row": row_idx, "values": [v for v in row[:6]]})
    else:
        # No header: treat each line as a bare cNumber
        for row in data_rows:
            cn = row[0].strip() if row else ""
            if not cn:
                continue
            entries.append({
                "id": str(uuid.uuid4()), "list_type": list_type,
                "entry_type": "cnumber", "cnumber": cn,
                "chrom": "", "pos": 0, "ref": "", "alt": "", "gene": "",
                "note": row[1].strip() if len(row) > 1 else "",
                "category": "", "ratio": "",
                "created_at": now, "updated_at": now,
            })

    return entries, skipped_rows


# ── API: Variant Lists CRUD ──────────────────────────────────────────────────

@app.route("/api/variant-lists/download/<list_type>")
def api_vl_download(list_type):
    """Download a variant list as xlsx in the standard upload format:
    Columns: Gene, Variant, 분류, 빈도, 비율
    """
    _require_auth()
    if list_type not in ("blacklist", "whitelist", "onhold"):
        return jsonify({"error": "invalid list_type"}), 400
    db = get_db()
    rows = db.execute(
        "SELECT entry_type, cnumber, chrom, pos, ref, alt, gene, category, ratio, note FROM variant_lists WHERE list_type=? ORDER BY created_at DESC",
        (list_type,)
    ).fetchall()

    try:
        import openpyxl
        from io import BytesIO
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = list_type.capitalize()
    headers = ["Gene", "Variant", "분류", "빈도", "비율", "비고"]
    ws.append(headers)

    for r in rows:
        r = dict_from_row(r)
        et = r.get("entry_type", "")
        gene = r.get("gene", "") or ""
        cat  = r.get("category", "") or ""
        ratio = r.get("ratio", "") or ""
        note = r.get("note", "") or ""

        if et == "gene":
            variant = ""
        elif et == "position":
            chrom = (r.get("chrom") or "").replace("chr", "")
            pos   = r.get("pos") or ""
            ref   = r.get("ref") or ""
            alt   = r.get("alt") or ""
            if chrom and pos:
                variant = f"{chrom}-{pos}-{ref}-{alt}" if ref and alt else f"{chrom}-{pos}"
            else:
                variant = r.get("cnumber") or ""
        else:  # cnumber
            variant = r.get("cnumber") or ""

        ws.append([gene, variant, cat, ratio, "", note])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{list_type}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


@app.route("/api/variant-lists")
def api_vl_list():
    """GET /api/variant-lists?type=blacklist|whitelist|onhold|longitudinal  (or all if omitted)."""
    list_type = request.args.get("type", "")
    db = get_db()
    if list_type in ("blacklist", "whitelist", "onhold"):
        rows = db.execute(
            "SELECT * FROM variant_lists WHERE list_type=? ORDER BY created_at DESC", (list_type,)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM variant_lists ORDER BY list_type, created_at DESC"
        ).fetchall()
    return jsonify([dict_from_row(r) for r in rows])


@app.route("/api/variant-lists", methods=["POST"])
def api_vl_create():
    """Add a single variant-list entry."""
    uid = session.get("user_id")
    if not uid:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    data = request.json or {}
    list_type = data.get("list_type", "")
    if list_type not in ("blacklist", "whitelist", "onhold"):
        return jsonify({"success": False, "error": "list_type must be blacklist/whitelist/onhold"}), 400
    entry_type = data.get("entry_type", "position")
    if entry_type not in ("cnumber", "position", "gene"):
        return jsonify({"success": False, "error": "entry_type must be cnumber/position/gene"}), 400

    # genome: 'any' for reference-independent types, else 'hg38'/'hg19'
    if entry_type in ("cnumber", "gene"):
        genome = "any"
    else:
        genome = data.get("genome", "hg38")
        if genome not in ("hg38", "hg19", "any"):
            genome = "hg38"

    now = datetime.now().isoformat()
    vid = str(uuid.uuid4())
    db = get_db()
    db.execute(
        """INSERT INTO variant_lists
           (id, list_type, entry_type, cnumber, chrom, pos, ref, alt, gene, note,
            category, ratio, genome, created_at, updated_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            vid, list_type, entry_type,
            (data.get("cnumber") or "").strip(),
            (data.get("chrom") or "").strip(),
            int(data.get("pos") or 0),
            (data.get("ref") or "").strip().upper(),
            (data.get("alt") or "").strip().upper(),
            (data.get("gene") or "").strip(),
            (data.get("note") or "").strip(),
            (data.get("category") or "").strip(),
            (data.get("ratio") or "").strip(),
            genome,
            now, now,
        ),
    )
    db.commit()
    return jsonify({"success": True, "id": vid})


@app.route("/api/variant-lists/<entry_id>", methods=["PUT"])
def api_vl_update(entry_id):
    uid = session.get("user_id")
    if not uid:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    db = get_db()
    row = db.execute("SELECT id FROM variant_lists WHERE id=?", (entry_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "error": "Not found"}), 404
    data = request.json or {}
    now = datetime.now().isoformat()
    editable = ["list_type", "entry_type", "cnumber", "chrom", "pos", "ref", "alt", "gene", "note", "category", "ratio", "genome"]
    sets = ["updated_at=?"]
    params = [now]
    for col in editable:
        if col in data:
            val = data[col]
            if col in ("ref", "alt") and isinstance(val, str):
                val = val.upper()
            if col == "pos":
                val = int(val or 0)
            if col == "genome" and val not in ("hg38", "hg19", "any"):
                val = "hg38"
            sets.append(f"{col}=?")
            params.append(val)
    params.append(entry_id)
    db.execute(f"UPDATE variant_lists SET {', '.join(sets)} WHERE id=?", params)
    db.commit()
    return jsonify({"success": True})


@app.route("/api/variant-lists/<entry_id>", methods=["DELETE"])
def api_vl_delete(entry_id):
    uid = session.get("user_id")
    if not uid:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    db = get_db()
    db.execute("DELETE FROM variant_lists WHERE id=?", (entry_id,))
    db.commit()
    return jsonify({"success": True})


@app.route("/api/variant-lists/upload", methods=["POST"])
def api_vl_upload():
    """Upload an xlsx / txt / bed file and bulk-insert entries."""
    uid = session.get("user_id")
    if not uid:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    list_type = request.form.get("list_type", "")
    if list_type not in ("blacklist", "whitelist", "onhold"):
        return jsonify({"success": False, "error": "list_type must be blacklist/whitelist/onhold"}), 400
    skip_header = request.form.get("skip_header", "false").lower() in ("true", "1", "yes")
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"success": False, "error": "파일이 전달되지 않았습니다."}), 400
    file_bytes = f.read()
    try:
        entries, skipped = _parse_variant_list_file(file_bytes, f.filename, list_type, skip_header=skip_header)
    except Exception as e:
        return jsonify({"success": False, "error": f"파일 파싱 오류: {e}"}), 400
    if not entries:
        return jsonify({"success": False, "error": "파싱된 항목이 없습니다. 파일 형식을 확인해주세요."}), 400
    # Ensure category/ratio keys exist (older entries from bed/txt may lack them)
    for e in entries:
        e.setdefault("category", "")
        e.setdefault("ratio", "")
    db = get_db()
    db.executemany(
        """INSERT OR IGNORE INTO variant_lists
           (id, list_type, entry_type, cnumber, chrom, pos, ref, alt, gene, note,
            category, ratio, created_at, updated_at)
           VALUES (:id,:list_type,:entry_type,:cnumber,:chrom,:pos,:ref,:alt,:gene,:note,
                   :category,:ratio,:created_at,:updated_at)""",
        entries,
    )
    db.commit()
    return jsonify({"success": True, "inserted": len(entries), "skipped": skipped})


# ---------------------------------------------------------------------------
# RNA-seq Interactive Analysis API
# ---------------------------------------------------------------------------

@app.route("/api/rna/gene_search")
def api_rna_gene_search():
    """Search for a gene in the expression summary of a sample."""
    sample = request.args.get("sample", "").strip()
    query = request.args.get("q", "").strip().upper()
    if not sample or not query:
        return jsonify({"error": "sample and q parameters required"}), 400

    expr_dir = os.path.join(RESULTS_DIR, sample, "expression_plots")
    summary_path = os.path.join(expr_dir, f"{sample}_expression_summary.tsv")
    if not os.path.isfile(summary_path):
        return jsonify({"error": "Expression summary not found"}), 404

    try:
        import csv as _csv
        results = []
        with open(summary_path, "r") as fh:
            reader = _csv.DictReader(fh, delimiter="\t")
            for row in reader:
                gene_id = (row.get("gene_id") or "").upper()
                gene_sym = (row.get("gene_symbol") or "").upper()
                if query in gene_id or query in gene_sym:
                    results.append({
                        "gene_id": row.get("gene_id", ""),
                        "gene_symbol": row.get("gene_symbol", ""),
                        "count": float(row.get("count", 0)),
                        "CPM": float(row.get("CPM", 0)),
                        "TPM": float(row.get("TPM", 0)),
                        "log2CPM": float(row.get("log2CPM", 0)),
                    })
                    if len(results) >= 50:
                        break
        return jsonify({"results": results, "total": len(results)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/rna/expression_data")
def api_rna_expression_data():
    """Return expression data for a specific gene across all samples or for a single sample."""
    sample = request.args.get("sample", "").strip()
    gene_id = request.args.get("gene_id", "").strip()
    gene_symbol = request.args.get("gene_symbol", "").strip()
    if not sample:
        return jsonify({"error": "sample parameter required"}), 400
    if not gene_id and not gene_symbol:
        return jsonify({"error": "gene_id or gene_symbol parameter required"}), 400

    expr_dir = os.path.join(RESULTS_DIR, sample, "expression_plots")
    summary_path = os.path.join(expr_dir, f"{sample}_expression_summary.tsv")
    if not os.path.isfile(summary_path):
        return jsonify({"error": "Expression summary not found"}), 404

    try:
        import csv as _csv
        gene_data = None
        with open(summary_path, "r") as fh:
            reader = _csv.DictReader(fh, delimiter="\t")
            for row in reader:
                row_gene_id = (row.get("gene_id") or "").upper()
                row_gene_sym = (row.get("gene_symbol") or "").upper()
                if (gene_id and gene_id.upper() in row_gene_id) or \
                   (gene_symbol and gene_symbol.upper() == row_gene_sym):
                    gene_data = {
                        "gene_id": row.get("gene_id", ""),
                        "gene_symbol": row.get("gene_symbol", ""),
                        "count": float(row.get("count", 0)),
                        "CPM": float(row.get("CPM", 0)),
                        "TPM": float(row.get("TPM", 0)),
                        "log2CPM": float(row.get("log2CPM", 0)),
                        "length": int(row.get("length", 0)),
                    }
                    break
        if gene_data is None:
            return jsonify({"error": "Gene not found"}), 404
        return jsonify({"gene": gene_data, "sample": sample})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/rna/gene_expression_across_samples")
def api_rna_gene_expression_across_samples():
    """
    Return expression values (TPM, CPM, count) for a specific gene
    across ALL completed RNA samples, enabling cross-sample comparison.
    """
    gene_query = request.args.get("gene", "").strip().upper()
    metric = request.args.get("metric", "TPM")  # TPM | CPM | count | log2CPM
    if not gene_query:
        return jsonify({"error": "gene parameter required"}), 400
    if metric not in ("TPM", "CPM", "count", "log2CPM"):
        metric = "TPM"

    # Find all expression summary files across all samples
    pattern = os.path.join(RESULTS_DIR, "*", "expression_plots", "*_expression_summary.tsv")
    summary_files = sorted(glob.glob(pattern))

    results = []
    for fpath in summary_files:
        parts = fpath.replace("\\", "/").split("/")
        # path: RESULTS_DIR / <sample> / expression_plots / <sample>_expression_summary.tsv
        sample_name = parts[-3]
        try:
            import csv as _csv
            with open(fpath, "r") as fh:
                reader = _csv.DictReader(fh, delimiter="\t")
                for row in reader:
                    gid = (row.get("gene_id") or "").upper()
                    gsym = (row.get("gene_symbol") or "").upper()
                    # Match by symbol (exact) or ID (contains)
                    if gene_query == gsym or gene_query in gid:
                        results.append({
                            "sample": sample_name,
                            "gene_id": row.get("gene_id", ""),
                            "gene_symbol": row.get("gene_symbol", ""),
                            "count": float(row.get("count", 0) or 0),
                            "CPM": float(row.get("CPM", 0) or 0),
                            "TPM": float(row.get("TPM", 0) or 0),
                            "log2CPM": float(row.get("log2CPM", 0) or 0),
                            "length": int(row.get("length", 0) or 0),
                        })
                        break  # one match per sample
        except Exception:
            continue

    if not results:
        return jsonify({"error": f"Gene '{gene_query}' not found in any sample", "samples": []}), 404

    # Sort by sample name
    results.sort(key=lambda x: x["sample"])
    return jsonify({
        "gene": results[0]["gene_symbol"] or results[0]["gene_id"],
        "gene_id": results[0]["gene_id"],
        "metric": metric,
        "samples": results,
        "total_samples": len(results),
    })


@app.route("/api/rna/multi_gene_expression")
def api_rna_multi_gene_expression():
    """
    Return expression values for multiple genes in a single sample.
    Used for gene group / pathway comparison.
    Query params:
      sample  : sample name
      genes   : comma-separated gene symbols or IDs
      metric  : TPM | CPM | count | log2CPM (default TPM)
    """
    sample = request.args.get("sample", "").strip()
    genes_raw = request.args.get("genes", "").strip()
    metric = request.args.get("metric", "TPM")
    if not sample or not genes_raw:
        return jsonify({"error": "sample and genes parameters required"}), 400
    if metric not in ("TPM", "CPM", "count", "log2CPM"):
        metric = "TPM"

    gene_list = [g.strip().upper() for g in genes_raw.split(",") if g.strip()]

    summary_path = os.path.join(RESULTS_DIR, sample, "expression_plots", f"{sample}_expression_summary.tsv")
    if not os.path.isfile(summary_path):
        return jsonify({"error": "Expression summary not found"}), 404

    try:
        import csv as _csv
        found = {}
        with open(summary_path, "r") as fh:
            reader = _csv.DictReader(fh, delimiter="\t")
            for row in reader:
                gid = (row.get("gene_id") or "").upper()
                gsym = (row.get("gene_symbol") or "").upper()
                for q in gene_list:
                    if q == gsym or q in gid:
                        found[q] = {
                            "query": q,
                            "gene_id": row.get("gene_id", ""),
                            "gene_symbol": row.get("gene_symbol", ""),
                            "count": float(row.get("count", 0) or 0),
                            "CPM": float(row.get("CPM", 0) or 0),
                            "TPM": float(row.get("TPM", 0) or 0),
                            "log2CPM": float(row.get("log2CPM", 0) or 0),
                        }
                        break

        # Fill missing genes with zeros
        results = []
        for q in gene_list:
            if q in found:
                results.append(found[q])
            else:
                results.append({"query": q, "gene_id": "", "gene_symbol": q,
                                 "count": 0, "CPM": 0, "TPM": 0, "log2CPM": 0})

        return jsonify({"sample": sample, "metric": metric, "genes": results})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/rna/all_samples")
def api_rna_all_samples():
    """Return list of all RNA samples that have expression summary files."""
    pattern = os.path.join(RESULTS_DIR, "*", "expression_plots", "*_expression_summary.tsv")
    files = sorted(glob.glob(pattern))
    samples = []
    for f in files:
        parts = f.replace("\\", "/").split("/")
        if len(parts) >= 3:
            sample_name = parts[-3]
            samples.append(sample_name)
    return jsonify({"samples": samples, "total": len(samples)})


@app.route("/api/rna/rseqc_data")
def api_rna_rseqc_data():
    """Return parsed RSeQC results for a sample."""
    sample = request.args.get("sample", "").strip()
    if not sample:
        return jsonify({"error": "sample parameter required"}), 400

    rseqc_dir = os.path.join(RESULTS_DIR, sample, "QC_report", "rseqc")
    result = {"sample": sample}

    # infer_experiment
    infer_path = os.path.join(rseqc_dir, f"{sample}_infer_experiment.txt")
    if os.path.isfile(infer_path):
        try:
            with open(infer_path, "r") as fh:
                result["infer_experiment"] = fh.read()
        except Exception:
            pass

    # read_distribution
    read_dist_path = os.path.join(rseqc_dir, f"{sample}_read_distribution.txt")
    if os.path.isfile(read_dist_path):
        try:
            with open(read_dist_path, "r") as fh:
                result["read_distribution"] = fh.read()
        except Exception:
            pass

    # TIN summary
    tin_candidates = glob.glob(os.path.join(rseqc_dir, "*.summary.txt"))
    if tin_candidates:
        try:
            with open(tin_candidates[0], "r") as fh:
                result["tin_summary"] = fh.read()
        except Exception:
            pass

    return jsonify(result)


# ---------------------------------------------------------------------------
# RNA-seq DESeq2 & Pathway Analysis API
# ---------------------------------------------------------------------------

DEG_DIR = os.path.join(RESULTS_DIR, "..", "Differential_Expression")  # relative to RESULTS_DIR
PATHWAY_DIR = os.path.join(RESULTS_DIR, "..", "Pathway_Analysis")


@app.route("/api/rna/deseq2_results")
def api_rna_deseq2_results():
    """Return DESeq2 differential expression results."""
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 100))
    sig_only = request.args.get("sig_only", "false").lower() == "true"
    sort_by = request.args.get("sort_by", "padj")  # padj | log2FoldChange | baseMean

    # Search in RESULTS_DIR parent (pipeline outdir)
    outdir = RESULTS_DIR
    deg_candidates = [
        os.path.join(outdir, "Differential_Expression", "deseq2_results.tsv"),
        os.path.join(os.path.dirname(outdir), "Differential_Expression", "deseq2_results.tsv"),
    ]
    results_path = next((p for p in deg_candidates if os.path.isfile(p)), None)

    if not results_path:
        return jsonify({"error": "DESeq2 results not found. Run pipeline with >= 2 samples."}), 404

    try:
        import csv as _csv
        rows = []
        with open(results_path, "r") as fh:
            reader = _csv.DictReader(fh, delimiter="\t")
            for row in reader:
                try:
                    row["padj"] = float(row.get("padj") or "nan")
                    row["log2FoldChange"] = float(row.get("log2FoldChange") or "nan")
                    row["baseMean"] = float(row.get("baseMean") or "nan")
                    row["pvalue"] = float(row.get("pvalue") or "nan")
                except ValueError:
                    pass
                rows.append(row)

        if sig_only:
            rows = [r for r in rows if isinstance(r.get("padj"), float)
                    and r["padj"] < 0.05
                    and abs(r.get("log2FoldChange", 0)) > 1]

        # Sort
        import math
        rows.sort(key=lambda r: (math.isnan(r.get(sort_by, float("nan"))) if isinstance(r.get(sort_by), float) else True,
                                  r.get(sort_by, float("inf"))))

        total = len(rows)
        start = (page - 1) * per_page
        end = start + per_page
        page_rows = rows[start:end]

        return jsonify({
            "total": total,
            "page": page,
            "per_page": per_page,
            "results": page_rows,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/rna/deseq2_plots")
def api_rna_deseq2_plots():
    """Return available DESeq2 plot file paths."""
    outdir = RESULTS_DIR
    deg_dir_candidates = [
        os.path.join(outdir, "Differential_Expression"),
        os.path.join(os.path.dirname(outdir), "Differential_Expression"),
    ]
    deg_dir = next((d for d in deg_dir_candidates if os.path.isdir(d)), None)
    if not deg_dir:
        return jsonify({"plots": [], "has_interactive": False})

    png_files = glob.glob(os.path.join(deg_dir, "*.png"))
    html_files = glob.glob(os.path.join(deg_dir, "*.html"))

    plots = [os.path.relpath(p, RESULTS_DIR) for p in sorted(png_files)]
    interactive = [os.path.relpath(p, RESULTS_DIR) for p in sorted(html_files)]

    return jsonify({
        "plots": plots,
        "interactive": interactive,
        "has_interactive": len(interactive) > 0,
        "deg_dir": deg_dir,
    })


@app.route("/api/rna/pathway_results")
def api_rna_pathway_results():
    """Return GO/KEGG pathway enrichment results."""
    pathway_type = request.args.get("type", "go")  # go | kegg
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 50))

    outdir = RESULTS_DIR
    pw_dir_candidates = [
        os.path.join(outdir, "Pathway_Analysis"),
        os.path.join(os.path.dirname(outdir), "Pathway_Analysis"),
    ]
    pw_dir = next((d for d in pw_dir_candidates if os.path.isdir(d)), None)

    if not pw_dir:
        return jsonify({"error": "Pathway analysis results not found."}), 404

    fname = "go_enrichment.tsv" if pathway_type == "go" else "kegg_enrichment.tsv"
    fpath = os.path.join(pw_dir, fname)

    if not os.path.isfile(fpath) or os.path.getsize(fpath) == 0:
        return jsonify({"results": [], "total": 0, "message": f"No {pathway_type.upper()} enrichment results."})

    try:
        import csv as _csv
        rows = []
        with open(fpath, "r") as fh:
            reader = _csv.DictReader(fh, delimiter="\t")
            for row in reader:
                try:
                    row["pvalue"] = float(row.get("pvalue") or "nan")
                    row["p.adjust"] = float(row.get("p.adjust") or "nan")
                    row["qvalue"] = float(row.get("qvalue") or "nan")
                except ValueError:
                    pass
                rows.append(row)

        total = len(rows)
        start = (page - 1) * per_page
        page_rows = rows[start:start + per_page]

        return jsonify({"total": total, "page": page, "per_page": per_page, "results": page_rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/rna/pathway_plots")
def api_rna_pathway_plots():
    """Return available pathway analysis plot file paths."""
    outdir = RESULTS_DIR
    pw_dir_candidates = [
        os.path.join(outdir, "Pathway_Analysis"),
        os.path.join(os.path.dirname(outdir), "Pathway_Analysis"),
    ]
    pw_dir = next((d for d in pw_dir_candidates if os.path.isdir(d)), None)
    if not pw_dir:
        return jsonify({"plots": []})

    png_files = glob.glob(os.path.join(pw_dir, "*.png"))
    plots = [os.path.relpath(p, RESULTS_DIR) for p in sorted(png_files)]
    return jsonify({"plots": plots})


@app.route("/api/rna/upload_design", methods=["POST"])
def api_rna_upload_design():
    """
    Upload a sample design CSV for DESeq2 group comparison.
    Format: sample,condition (header required)
    Saves to RESULTS_DIR/design.csv
    """
    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file provided"}), 400
    f = request.files["file"]
    if not f.filename.endswith(".csv"):
        return jsonify({"success": False, "error": "Only CSV files accepted"}), 400
    try:
        content = f.read().decode("utf-8")
        lines = [l.strip() for l in content.splitlines() if l.strip()]
        if not lines or "sample" not in lines[0].lower():
            return jsonify({"success": False, "error": "CSV must have header: sample,condition"}), 400
        dest = os.path.join(RESULTS_DIR, "design.csv")
        with open(dest, "w") as fh:
            fh.write(content)
        return jsonify({"success": True, "message": f"Design saved to {dest}", "rows": len(lines) - 1})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
init_db()

if __name__ == "__main__":
    for d in [FASTQ_DIR, RESULTS_DIR, LOG_DIR]:
        os.makedirs(d, exist_ok=True)
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
